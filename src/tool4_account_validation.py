# -*- coding: utf-8 -*-
"""
ツール④ 口座データバリデーション（OCR転記支援）
生成ファイル: output/口座データバリデーション.xlsx

シート構成:
  1. 口座データ入力    : 紙の申込書から転記。水色セルに入力
  2. バリデーション結果: 記号/番号/カナの自動チェック。K列「総合判定」で一目確認
  3. 銀行提出用        : 全行OKのデータを16桁コードに自動変換
  4. 操作ガイド        : 使い方・エラー対処法
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import sys, os

sys.path.insert(0, os.path.dirname(__file__))
from styles import (
    fill, font, border, align,
    style_header, style_input, style_readonly, style_formula,
    style_title, style_guide_row, setup_sheet, set_col_width, set_row_height,
    COLORS,
)

SAMPLE_ACCOUNTS = [
    # 生徒番号, 申込者氏名, 記号(5桁), 番号(8桁), 口座名義(半角カナ)
    ("04001", "山田 太郎",  "10050", "18233211", "ｱﾏﾀ ﾀﾛｳ"),
    ("04002", "鈴木 花子",  "10000", "08774961", "ｽｽﾞｷ ﾊﾅｺ"),
    ("04003", "田中 一郎",  "10280", "60231421", "ﾀﾅｶ ｲﾁﾛｳ"),
    ("04004", "佐藤 美咲",  "10150", "4567890",  "ｻﾄｳ ﾐｻｷ"),     # ← 番号7桁エラー
    ("04005", "高橋 健一",  "10740", "12345678", "タカハシ ケンイチ"),  # ← 全角カナエラー
    ("05001", "伊藤 奈々",  "10320", "87654321", "ｲﾄｳ ﾅﾅ"),
    ("05002", "中村 大輔",  "10560", "11223344", "ﾅｶﾑﾗ ﾀﾞｲｽｹ"),
    ("05003", "小林 あい",  "1089",  "55667788", "ｺﾊﾞﾔｼ ｱｲ"),     # ← 記号4桁エラー
    ("06001", "加藤 翼",    "10210", "99887766", "ｶﾄｳ ﾂﾊﾞｻ"),
    ("06002", "吉田 愛",    "10670", "44332211", "ﾖｼﾀﾞ ｱｲ"),
]

DATA_START = 4  # 入力シートのデータ開始行


def build_input_sheet(ws):
    ws.tab_color = COLORS["green"]
    style_title(ws, "口座データ入力 ─ 水色のセルだけに入力してください（紙の申込書から転記）",
                row=1, merge_cols=6, size=12)

    note = (
        "★ D列・E列（記号・番号）の書式は「文字列」です。先頭の0が消えないよう入力前に確認してください。\n"
        "★ F列（口座名義）は必ず半角カタカナで入力。全角「ア」でなく半角「ｱ」を使ってください。"
    )
    ws.cell(row=2, column=1).value = note
    ws.cell(row=2, column=1).font = font(size=9, color="red")
    ws.cell(row=2, column=1).alignment = align("left", "center", wrap=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 36

    col_headers = [
        "生徒番号\n(5桁・文字列)", "申込者氏名", "申込者ふりがな",
        "記号\n(5桁・数字)", "番号\n(8桁・数字)", "口座名義\n(半角カタカナ)",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=DATA_START - 1, column=col), h,
                     color="header_light", font_color="black", size=9)

    for i, row_data in enumerate(SAMPLE_ACCOUNTS, start=DATA_START):
        sid, name, kigo, bangou, meigi = row_data
        vals = [sid, name, "", kigo, bangou, meigi]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col)
            c.value = v
            if col in (1, 4, 5):
                c.number_format = "@"
            style_input(c)

    # 空白テンプレート
    for extra_r in range(DATA_START + len(SAMPLE_ACCOUNTS), DATA_START + 301):
        for col in range(1, 7):
            c = ws.cell(row=extra_r, column=col)
            if col in (1, 4, 5):
                c.number_format = "@"
            style_input(c)

    set_col_width(ws, {1: 14, 2: 16, 3: 16, 4: 12, 5: 14, 6: 24})
    set_row_height(ws, {1: 22, 2: 36, DATA_START - 1: 32})
    setup_sheet(ws, freeze_row=DATA_START, freeze_col=2)


def build_validation_sheet(ws):
    ws.tab_color = COLORS["red"]
    style_title(ws, "バリデーション結果 ─ K列「総合判定」がすべて OK になるまで修正してください",
                row=1, merge_cols=11, size=12)

    # 凡例
    ws.cell(row=2, column=1).value = (
        "【判定の見方】 OK（緑）→ 銀行提出可　要確認（赤）→ エラーあり。"
        "G〜J列で原因を確認して「口座データ入力」シートを修正してください。"
    )
    ws.cell(row=2, column=1).font = font(size=9)
    ws.cell(row=2, column=1).alignment = align("left", "center", wrap=True)
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 28

    col_headers = [
        "生徒番号", "氏名", "記号(参考)", "番号(参考)", "名義(参考)",
        "記号チェック", "番号チェック", "カナチェック", "重複チェック",
        "総合判定", "修正メモ",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=3, column=col), h,
                     color="header_light", font_color="black", size=9)

    # CoPilot向けメタデータコメント
    ws.cell(row=3, column=6).comment = None  # placeholder for openpyxl
    ws.cell(row=3, column=10).fill = fill("header")
    ws.cell(row=3, column=10).font = font(bold=True, color="white", size=9)

    SRC = "口座データ入力"
    for i in range(DATA_START, DATA_START + 302):
        src_r = i  # 口座データ入力シートの行番号は同じ

        # A: 生徒番号
        c = ws.cell(row=i, column=1)
        c.value = f"=IF('{SRC}'!A{src_r}=\"\",\"\",'{SRC}'!A{src_r})"
        c.number_format = "@"
        style_formula(c)

        # B: 氏名
        c = ws.cell(row=i, column=2)
        c.value = (
            f"=IF(A{i}=\"\",\"\","
            f"IFERROR(XLOOKUP(A{i},生徒マスター!$A:$A,生徒マスター!$B:$B),\"未マッチ\"))"
        )
        style_formula(c)

        # C: 記号 参考表示
        c = ws.cell(row=i, column=3)
        c.value = f"=IF('{SRC}'!D{src_r}=\"\",\"\",'{SRC}'!D{src_r})"
        c.number_format = "@"
        style_formula(c)

        # D: 番号 参考表示
        c = ws.cell(row=i, column=4)
        c.value = f"=IF('{SRC}'!E{src_r}=\"\",\"\",'{SRC}'!E{src_r})"
        c.number_format = "@"
        style_formula(c)

        # E: 名義 参考表示
        c = ws.cell(row=i, column=5)
        c.value = f"=IF('{SRC}'!F{src_r}=\"\",\"\",'{SRC}'!F{src_r})"
        style_formula(c)

        # F: 記号チェック（5桁かつ数字のみ）
        c = ws.cell(row=i, column=6)
        c.value = (
            f'=IF(C{i}="","未入力",'
            f'IF(AND(LEN(C{i})=5,ISNUMBER(VALUE(C{i}))),"OK","NG: "&LEN(C{i})&"桁"))'
        )
        style_formula(c)

        # G: 番号チェック（8桁かつ数字のみ）
        c = ws.cell(row=i, column=7)
        c.value = (
            f'=IF(D{i}="","未入力",'
            f'IF(AND(LEN(D{i})=8,ISNUMBER(VALUE(D{i}))),"OK","NG: "&LEN(D{i})&"桁"))'
        )
        style_formula(c)

        # H: カナチェック（半角カナ = LENB=LEN）
        c = ws.cell(row=i, column=8)
        c.value = (
            f'=IF(E{i}="","未入力",'
            f'IF(LENB(E{i})=LEN(E{i}),"OK","NG: 全角混在"))'
        )
        style_formula(c)

        # I: 重複チェック
        c = ws.cell(row=i, column=9)
        c.value = (
            f'=IF(OR(C{i}="",D{i}=""),"─",'
            f"IF(COUNTIFS('{SRC}'!$D:$D,C{i},'{SRC}'!$E:$E,D{i})>1,"
            f'"重複あり","OK"))'
        )
        style_formula(c)

        # J: 総合判定
        c = ws.cell(row=i, column=10)
        c.value = (
            f'=IF(A{i}="","─",'
            f'IF(AND(F{i}="OK",G{i}="OK",H{i}="OK",I{i}="OK"),"OK","要確認"))'
        )
        style_formula(c)
        c.font = font(bold=True, size=10)
        c.alignment = align("center", "center")

        # K: 修正メモ（手入力欄）
        style_input(ws.cell(row=i, column=11))

    # ── 条件付き書式 ──────────────────────────────────────────────────────
    data_range = f"A{DATA_START}:K{DATA_START + 301}"

    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'$J{DATA_START}="OK"'], fill=fill("green_light")),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'$J{DATA_START}="要確認"'], fill=fill("red_light")),
    )

    # チェック列 (F〜J) の個別色
    check_range = f"F{DATA_START}:J{DATA_START + 301}"
    ws.conditional_formatting.add(
        check_range,
        FormulaRule(formula=[f'LEFT(F{DATA_START},2)="NG"'], fill=fill("orange_light")),
    )

    set_col_width(ws, {
        1: 12, 2: 16, 3: 10, 4: 12, 5: 22,
        6: 14, 7: 14, 8: 14, 9: 12, 10: 12, 11: 20,
    })
    set_row_height(ws, {1: 22, 2: 28, 3: 32})
    setup_sheet(ws, freeze_row=DATA_START, freeze_col=2)


def build_bank_sheet(ws):
    ws.tab_color = COLORS["navy"]
    style_title(ws, "銀行提出用 ─ 総合判定OKの行のみ自動生成。そのまま銀行に提出してください",
                row=1, merge_cols=5, size=12)

    ws.cell(row=2, column=1).value = (
        "⚠ 16桁コードの書式は銀行仕様書に依存します。提出前に銀行担当者と書式を確認してください。\n"
        "「─」と表示されている行はバリデーションエラーのため自動除外されています。"
    )
    ws.cell(row=2, column=1).font = font(size=9, color="red")
    ws.cell(row=2, column=1).alignment = align("left", "center", wrap=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 36

    col_headers = [
        "生徒番号", "氏名", "記号", "番号", "16桁コード\n(要銀行仕様確認)",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=3, column=col), h,
                     color="navy", font_color="white", size=9)

    VAL = "バリデーション結果"
    SRC = "口座データ入力"
    for i in range(DATA_START, DATA_START + 302):
        # A: 生徒番号
        c = ws.cell(row=i, column=1)
        c.value = f"=IF('{VAL}'!J{i}=\"OK\",'{VAL}'!A{i},\"─\")"
        c.number_format = "@"
        style_formula(c)
        c.alignment = align("center")

        # B: 氏名
        c = ws.cell(row=i, column=2)
        c.value = f"=IF(A{i}=\"─\",\"─\",'{VAL}'!B{i})"
        style_formula(c)

        # C: 記号
        c = ws.cell(row=i, column=3)
        c.value = f"=IF(A{i}=\"─\",\"─\",'{SRC}'!D{i})"
        c.number_format = "@"
        style_formula(c)
        c.alignment = align("center")

        # D: 番号
        c = ws.cell(row=i, column=4)
        c.value = f"=IF(A{i}=\"─\",\"─\",'{SRC}'!E{i})"
        c.number_format = "@"
        style_formula(c)
        c.alignment = align("center")

        # E: 16桁コード（生徒番号0埋め8桁 + 記号5桁 + 固定"000"）
        # ※実際の銀行仕様に合わせてこの数式を修正してください
        c = ws.cell(row=i, column=5)
        c.value = (
            f'=IF(A{i}="─","─",'
            f'TEXT(A{i},"00000000")&TEXT(\'{SRC}\'!D{i},"00000")&"000")'
        )
        c.number_format = "@"
        style_formula(c)
        c.font = font(size=10, name="Consolas")
        c.alignment = align("center")

    # 集計
    SUMMARY_ROW = DATA_START + 302
    ws.cell(row=SUMMARY_ROW, column=1).value = "提出件数"
    ws.cell(row=SUMMARY_ROW, column=1).font = font(bold=True)
    ws.cell(row=SUMMARY_ROW, column=1).border = border()
    c = ws.cell(row=SUMMARY_ROW, column=2)
    c.value = f'=COUNTIF(A{DATA_START}:A{SUMMARY_ROW-1},"<>─")-COUNTBLANK(A{DATA_START}:A{SUMMARY_ROW-1})'
    c.font = font(bold=True, size=12)
    c.border = border()
    c.number_format = "0\"件\""
    c.alignment = align("center")

    set_col_width(ws, {1: 12, 2: 16, 3: 10, 4: 14, 5: 22})
    set_row_height(ws, {1: 22, 2: 36, 3: 32})
    setup_sheet(ws, freeze_row=DATA_START, freeze_col=2)


def build_guide_sheet(ws):
    ws.tab_color = COLORS["header"]
    style_title(ws, "操作ガイド ─ 口座データバリデーション（自動払込利用申込書の処理）",
                row=1, merge_cols=4, size=13)

    rows = [
        ("【このツールの目的】",
         "保護者から届いた紙の自動払込利用申込書をExcelに入力した後、\n"
         "記号の桁数・番号の桁数・カナの全角/半角ミスを自動チェックします。\n"
         "3名で目視チェックしていた1〜2日の作業が30分程度になります。"),
        ("【手順①：入力】",
         "「口座データ入力」シート（緑タブ）を開く\n"
         "→ 水色のセルに紙の申込書を見ながら入力（または貼り付け）\n"
         "→ D列・E列は「文字列」書式。先頭の0が消えないか確認。\n"
         "→ F列（口座名義）は半角カタカナで入力。全角カタカナは不可。"),
        ("【手順②：チェック】",
         "「バリデーション結果」シート（赤タブ）を開く\n"
         "→ K列「総合判定」を確認。「要確認」の行だけを修正する。\n"
         "→ G〜J列で具体的なエラー内容を確認。"),
        ("【手順③：提出】",
         "全行が「OK」になったら「銀行提出用」シート（紺タブ）を開く\n"
         "→ 16桁コードが自動生成されています。\n"
         "→ ⚠ コードの書式は事前に銀行担当者と確認してください。"),
        ("【エラー対処】",
         "NG: 4桁 → 記号が5桁でない。紙を見て正しい5桁を入力し直す\n"
         "NG: 7桁 → 番号の先頭0が消えた。セルを文字列書式にして0付きで入力し直す\n"
         "全角混在 → 全角カタカナが混じっている。半角カタカナに直す\n"
         "重複あり → 同じ口座が2行以上ある。紙を確認して正しい方を残す\n"
         "未マッチ → 申込者が生徒マスターにいない（転入生 or 名前の表記ゆれ）"),
        ("⚠ 先頭0消失の防止",
         "D列（記号）・E列（番号）を入力前に「文字列」書式に設定してください。\n"
         "方法: D列を選択 → 右クリック → セルの書式設定 → 文字列 → OK\n"
         "テンプレートではあらかじめ設定済みですが、コピー&ペーストで崩れる場合があります。"),
        ("✨ CoPilotの活用",
         "このシートをCoPilotに渡して「エラー行を一覧にして」と聞くと、\n"
         "要確認行だけを抽出して修正ポイントを教えてくれます。\n"
         "例: 「このExcelのバリデーション結果シートで要確認になっている行を\n"
         "    すべてリストアップして、エラーの原因を教えてください」"),
    ]
    for i, (label, content) in enumerate(rows, start=3):
        style_guide_row(ws, i, label, content)
        ws.row_dimensions[i].height = 60

    set_col_width(ws, {"A": 22, "B": 72})
    setup_sheet(ws, freeze_row=3)


def create_workbook(output_path: str):
    wb = Workbook()

    ws_input = wb.active
    ws_input.title = "口座データ入力"
    build_input_sheet(ws_input)

    ws_val = wb.create_sheet("バリデーション結果")
    build_validation_sheet(ws_val)

    ws_bank = wb.create_sheet("銀行提出用")
    build_bank_sheet(ws_bank)

    ws_guide = wb.create_sheet("操作ガイド")
    build_guide_sheet(ws_guide)

    wb.save(output_path)
    print(f"  ✓ 口座データバリデーション.xlsx → {output_path}")


if __name__ == "__main__":
    os.makedirs("../output", exist_ok=True)
    create_workbook("../output/口座データバリデーション.xlsx")
