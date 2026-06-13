# -*- coding: utf-8 -*-
"""
ツール③ マスターブック（帳票自動生成）
生成ファイル: output/master_マスターブック.xlsx

シート構成:
  1. 操作ガイド        (説明書)
  2. 生徒マスター      (入力: 生徒情報・口座・フラグ)
  3. 支出項目マスター  (入力: 支出番号・金額・給付型フラグ)
  4. 月次徴収          (入力: 各回の請求/充当)
  5. 支出記録          (入力: 支出発生のたびに1行追記)
  6. 個人別管理表(出力)(自動: 118列形式を数式で再構成)
  7. 精算書(出力)      (自動: 生徒番号を入力→精算書生成)
  8. 決算書(出力)      (自動: 年度末決算書)
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys, os

sys.path.insert(0, os.path.dirname(__file__))
from styles import (
    fill, font, border, align,
    style_header, style_input, style_readonly, style_formula,
    style_title, style_guide_row, setup_sheet, set_col_width, set_row_height,
    COLORS,
)

# ── サンプルデータ ─────────────────────────────────────────────────────────────

STUDENTS = [
    # 生徒番号, 氏名,        クラス, ふりがな,          保護者氏名,      保護者ふりがな,       口座記号, 口座番号,   口座名義,         P, 給付型, 在籍状態, 前年度繰越
    ("04001","山田 太郎",  "Ｃ","やまだ たろう",   "山田 健一","やまだ けんいち","10050","18233211","ｱﾏﾀ ﾀﾛｳ",     "○","",    "在籍",28000),
    ("04002","鈴木 花子",  "Ａ","すずき はなこ",   "鈴木 一郎","すずき いちろう","10000","08774961","ｽｽﾞｷ ﾊﾅｺ",     "○","*(1)","在籍",31500),
    ("04003","田中 一郎",  "Ｂ","たなか いちろう", "田中 二郎","たなか じろう",  "10280","60231421","ﾀﾅｶ ｲﾁﾛｳ",     "×","",    "在籍",25000),
    ("04004","佐藤 美咲",  "Ｄ","さとう みさき",   "佐藤 良子","さとう よしこ",  "10150","45678901","ｻﾄｳ ﾐｻｷ",      "○","",    "在籍",0),
    ("04005","高橋 健一",  "Ｅ","たかはし けんいち","高橋 武", "たかはし たけし","10740","12345678","ﾀｶﾊｼ ｹﾝｲﾁ",    "○","*(2)","在籍",15000),
    ("05001","伊藤 奈々",  "Ａ","いとう なな",     "伊藤 裕子","いとう ゆうこ",  "10320","87654321","ｲﾄｳ ﾅﾅ",       "○","",    "在籍",0),
    ("05002","中村 大輔",  "Ｂ","なかむら だいすけ","中村 博", "なかむら ひろし","10560","11223344","ﾅｶﾑﾗ ﾀﾞｲｽｹ",   "○","",    "在籍",0),
    ("05003","小林 あい",  "Ｃ","こばやし あい",   "小林 幸子","こばやし さちこ","10890","55667788","ｺﾊﾞﾔｼ ｱｲ",     "○","*(1)","在籍",0),
    ("06001","加藤 翼",    "Ａ","かとう つばさ",   "加藤 隆", "かとう たかし",  "10210","99887766","ｶﾄｳ ﾂﾊﾞｻ",     "○","",    "在籍",0),
    ("06002","吉田 愛",    "Ｂ","よしだ あい",     "吉田 恵子","よしだ けいこ",  "10670","44332211","ﾖｼﾀﾞ ｱｲ",      "×","",    "在籍",0),
]

EXPENSE_ITEMS = [
    # 番号, 項目名,                    教科/区分,    単価, 対象人数, 執行時期,  給付型フラグ
    ( 1, "各教科フラットファイル",     "学校全体",    500,  190, "4月",   ""),
    ( 2, "芸術鑑賞教室（1年）",        "文化部",     1696,  190, "9月",   "*(1)"),
    ( 3, "芸術鑑賞教室（2年）",        "文化部",     2000,  190, "9月",   "*(1)"),
    ( 4, "修学旅行積立①",             "学校全体",  10000,  190, "5月",   "*(2)"),
    ( 5, "修学旅行積立②",             "学校全体",  10000,  190, "7月",   "*(2)"),
    ( 6, "修学旅行積立③",             "学校全体",  10000,  190, "9月",   "*(2)"),
    ( 7, "修学旅行積立④",             "学校全体",  10000,  190, "11月",  "*(2)"),
    ( 8, "修学旅行積立⑤",             "学校全体",  10000,  190, "1月",   "*(2)"),
    ( 9, "卒業アルバム",               "学校全体",   8000,  190, "3月",   ""),
    (10, "卒業式費用",                 "学校全体",   5000,  190, "3月",   ""),
    (11, "体育祭費",                   "体育",        300,  190, "5月",   ""),
    (12, "文化祭費",                   "文化部",      300,  190, "9月",   ""),
    (13, "球技大会費",                 "体育",        200,  190, "6月",   ""),
    (14, "生徒会費",                   "生徒会",     3000,  190, "4月",   ""),
    (15, "英語検定受験料",             "英語科",     2500,  100, "10月",  ""),
    (16, "漢字検定受験料",             "国語科",     1500,   80, "11月",  ""),
    (17, "情報処理検定受験料（3級）",  "商業",       1500,   60, "1月",   ""),
    (18, "情報処理検定受験料（2級）",  "商業",       2500,   40, "1月",   ""),
    (19, "簿記検定受験料",             "商業",       1500,   50, "2月",   ""),
    (20, "PTA会費",                    "PTA",        3000,  190, "4月",   ""),
    (21, "各教科副教材費（国語）",     "国語科",      800,  190, "4月",   "*(1)"),
    (22, "各教科副教材費（数学）",     "数学科",      600,  190, "4月",   "*(1)"),
    (23, "各教科副教材費（英語）",     "英語科",     1200,  190, "4月",   "*(1)"),
    (24, "各教科副教材費（理科）",     "理科",        700,  190, "4月",   "*(1)"),
    (25, "各教科副教材費（社会）",     "社会科",      500,  190, "4月",   "*(1)"),
    (26, "スポーツ振興費",             "体育",        500,  190, "4月",   ""),
    (27, "図書費",                     "図書",        800,  190, "4月",   "*(1)"),
    (28, "進路指導費",                 "進路",       1000,  190, "6月",   ""),
    (29, "校外学習費（1年）",          "学校全体",   3000,  190, "6月",   "*(2)"),
    (30, "校外学習費（2年）",          "学校全体",   3500,  190, "6月",   "*(2)"),
]

# 月次徴収サンプル（生徒番号, 1回請求, 1回充当, 2回請求, 2回充当, 3回請求, 3回充当, 4回請求, 4回充当, 5回請求, 5回充当, 現金振込）
MONTHLY = [
    ("04001", 35000,35000, 35000,35000, 35000,35000, 35000,35000, 35000,35000, 0),
    ("04002", 35000,35000, 35000,35000, 35000,35000, 35000,35000, 35000,35000, 0),
    ("04003", 35000,35000, 35000,35000, 35000,35000, 35000,35000, 35000,35000, 0),
    ("04004", 35000,35000, 35000,35000, 35000,35000, 35000,35000, 35000,35000, 0),
    ("04005", 35000,35000, 35000,35000, 35000,35000, 35000,35000, 35000,35000, 0),
    ("05001", 30000,30000, 30000,30000, 30000,30000, 30000,30000, 30000,30000, 0),
    ("05002", 30000,30000, 30000,30000, 30000,30000, 30000,30000, 30000,30000, 0),
    ("05003", 30000,30000, 30000,30000, 30000,30000, 30000,30000, 30000,30000, 0),
    ("06001", 25000,25000, 25000,25000, 25000,25000, 0,0, 0,0, 0),
    ("06002", 25000,25000, 25000,25000, 25000,25000, 0,0, 0,0, 0),
]

# 支出記録サンプル（日付, 支出番号, 生徒番号(空=全員), 金額, 備考）
EXPENSE_RECORDS = [
    ("2025/4/10",  1,  "",       500,  "全員分フラットファイル"),
    ("2025/4/10", 14,  "",      3000,  "生徒会費 全員"),
    ("2025/4/10", 20,  "",      3000,  "PTA会費 全員"),
    ("2025/9/15",  2,  "",      1696,  "芸術鑑賞教室"),
    ("2025/9/15",  2, "04002",  -1696, "欠席による返還"),
    ("2025/5/10",  4,  "",     10000,  "修学旅行積立①"),
    ("2025/7/10",  5,  "",     10000,  "修学旅行積立②"),
    ("2025/5/20", 11,  "",       300,  "体育祭費"),
    ("2025/9/20", 12,  "",       300,  "文化祭費"),
    ("2025/10/1", 15, "",       2500,  "英語検定"),
]

# 生徒マスターの列マッピング (1-indexed)
SM = {
    "生徒番号":   "A",
    "氏名":       "B",
    "クラス":     "C",
    "ふりがな":   "D",
    "保護者氏名": "E",
    "保護者かな": "F",
    "口座記号":   "G",
    "口座番号":   "H",
    "口座名義":   "I",
    "Pフラグ":    "J",
    "給付型":     "K",
    "在籍状態":   "L",
    "前年度繰越": "M",
}
SM_DATA_START = 4  # データ開始行

# 月次徴収の列マッピング
MT = {
    "生徒番号":   "A",
    "氏名":       "B",
    "クラス":     "C",
    "1回請求":    "D",
    "1回充当":    "E",
    "2回請求":    "F",
    "2回充当":    "G",
    "3回請求":    "H",
    "3回充当":    "I",
    "4回請求":    "J",
    "4回充当":    "K",
    "5回請求":    "L",
    "5回充当":    "M",
    "現金振込":   "N",
    "年間合計":   "O",
}
MT_DATA_START = 4


def build_guide_sheet(ws):
    ws.tab_color = COLORS["header"]
    style_title(ws, "操作ガイド ─ マスターブック 学次会計 帳票自動生成システム",
                row=1, merge_cols=4, size=13)

    rows = [
        ("【入力シートは3つだけ】",
         "① 生徒マスター（緑）　② 月次徴収（青）　③ 支出記録（赤）\n"
         "この3シートに入力するだけで、個人別管理表・精算書・決算書がすべて自動生成されます。"),
        ("【日常操作：支出発生時】",
         "「支出記録」シート（赤タブ）→ 一番下の空行に日付・支出番号・金額を入力。\n"
         "支出番号は「支出項目マスター」シートのA列に記載されています。\n"
         "全員同額の場合はD列（生徒番号）を空欄のまま。特定生徒のみはD列に生徒番号を入力。"),
        ("【月次操作：口座引落後】",
         "「月次徴収」シート（青タブ）→ 該当生徒の行に充当額を入力（充当はマイナスで入力）。"),
        ("【年度末：精算書を作る】",
         "「精算書(出力)」シート → B3の黄色セルに生徒番号を入力 → 自動表示。\n"
         "一括印刷は「マクロ実行」ボタンを押すか、VBAの「精算書_一括印刷」を実行。"),
        ("【年度更新：クラス替え後】",
         "「番号紐付けテンプレート.xlsx」でクラス照合 → 生徒マスターのC列（クラス）を更新。"),
        ("⚠ 給付型奨学金の自動処理",
         "生徒マスターのK列（給付型フラグ）が空欄でなく、かつ支出項目マスターのG列（給付型対象）も\n"
         "空欄でない場合、個人別管理表に自動で「給」と表示されます。2倍作成は不要です。"),
        ("⚠ セルの書式設定について",
         "生徒番号・口座記号・口座番号は「文字列」書式です。先頭の0が消えないよう\n"
         "コピー貼り付け後も書式が「文字列」になっているか確認してください。"),
        ("✨ 困ったときは",
         "・数式が#REF!などになった → 入力シートのデータが正しいか確認。出力シートは触らない。\n"
         "・自動計算が止まった → F9キーを押すか「数式」タブ→「計算方法」→「自動」\n"
         "・マクロを有効にするかどうか聞かれた → 「コンテンツの有効化」をクリック"),
    ]
    for i, (label, content) in enumerate(rows, start=3):
        style_guide_row(ws, i, label, content)
        ws.row_dimensions[i].height = 52

    set_col_width(ws, {"A": 26, "B": 75})
    setup_sheet(ws, freeze_row=3, freeze_col=1)


def build_student_master_sheet(ws):
    ws.tab_color = COLORS["green"]
    style_title(ws, "生徒マスター ─ 生徒情報・口座・フラグの一元管理（主キー: 生徒番号）",
                row=1, merge_cols=13, size=12)

    # セクションヘッダー（行2）
    sections = [
        (1,  "基本情報",   4),
        (5,  "保護者情報", 2),
        (7,  "口座情報",   3),
        (10, "フラグ",     2),
        (12, "状態・繰越", 2),
    ]
    for start, label, span in sections:
        c = ws.cell(row=2, column=start)
        c.value = label
        c.fill = fill("header_mid")
        c.font = font(bold=True, size=9, color="white")
        c.alignment = align("center")
        c.border = border()
        if span > 1:
            ws.merge_cells(start_row=2, start_column=start,
                           end_row=2, end_column=start + span - 1)

    # 列ヘッダー（行3）
    col_headers = [
        "生徒番号", "氏名", "クラス", "ふりがな",
        "保護者氏名", "保護者ふりがな",
        "口座記号\n(5桁)", "口座番号\n(8桁)", "口座名義\n(半角カナ)",
        "PTA\nフラグ", "給付型\nフラグ",
        "在籍状態", "前年度\n繰越金",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=3, column=col), h, color="header_light",
                     font_color="black", bold=True, size=9)

    # データ行
    for i, s in enumerate(STUDENTS, start=SM_DATA_START):
        vals = [s[0], s[1], s[2], s[3], s[4], s[5],
                s[6], s[7], s[8], s[9], s[10], s[11], s[12]]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col)
            c.value = v
            # 文字列書式が必要な列
            if col in (1, 7, 8):
                c.number_format = "@"
            # 前年度繰越金は数値書式
            if col == 13 and isinstance(v, (int, float)):
                c.number_format = "#,##0"
            style_input(c)

    # 空白テンプレート行（300名まで対応）
    for extra_r in range(SM_DATA_START + len(STUDENTS), SM_DATA_START + 301):
        for col in range(1, 14):
            c = ws.cell(row=extra_r, column=col)
            if col in (1, 7, 8):
                c.number_format = "@"
            style_input(c)

    # データ検証
    dv_class = DataValidation(type="list", formula1='"Ａ,Ｂ,Ｃ,Ｄ,Ｅ,Ｆ"', showDropDown=False)
    dv_class.sqref = f"C{SM_DATA_START}:C{SM_DATA_START+300}"
    ws.add_data_validation(dv_class)

    dv_pflag = DataValidation(type="list", formula1='"○,×"', showDropDown=False)
    dv_pflag.sqref = f"J{SM_DATA_START}:J{SM_DATA_START+300}"
    ws.add_data_validation(dv_pflag)

    dv_kyufutype = DataValidation(type="list", formula1='"*(1),*(2),*(3)"', showDropDown=False)
    dv_kyufutype.sqref = f"K{SM_DATA_START}:K{SM_DATA_START+300}"
    ws.add_data_validation(dv_kyufutype)

    dv_status = DataValidation(type="list", formula1='"在籍,転退学,卒業,原留"', showDropDown=False)
    dv_status.sqref = f"L{SM_DATA_START}:L{SM_DATA_START+300}"
    ws.add_data_validation(dv_status)

    set_col_width(ws, {
        1: 10, 2: 16, 3: 8, 4: 18,
        5: 16, 6: 18,
        7: 10, 8: 12, 9: 20,
        10: 7, 11: 9,
        12: 9, 13: 12,
    })
    set_row_height(ws, {1: 22, 2: 16, 3: 32})
    setup_sheet(ws, freeze_row=SM_DATA_START, freeze_col=3)


def build_expense_master_sheet(ws):
    ws.tab_color = COLORS["orange"]
    style_title(ws, "支出項目マスター ─ 支出番号・金額・給付型フラグの一覧",
                row=1, merge_cols=8, size=12)

    col_headers = [
        "支出番号", "項目名", "教科・区分",
        "単価(円)", "対象人数", "執行時期",
        "給付型\n対象フラグ", "合計金額\n(自動)",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=2, column=col), h, color="header_light",
                     font_color="black", bold=True, size=9)

    for i, item in enumerate(EXPENSE_ITEMS, start=3):
        num, name, category, price, count, timing, kyufu = item
        vals = [num, name, category, price, count, timing, kyufu]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col)
            c.value = v
            if col in (4, 5):
                c.number_format = "#,##0"
            style_input(c)
        # 合計金額 (数式)
        c = ws.cell(row=i, column=8)
        c.value = f"=IF(D{i}=\"\",\"\",D{i}*E{i})"
        c.number_format = "#,##0"
        style_formula(c)

    # 空白行（63項目まで対応）
    for extra_r in range(3 + len(EXPENSE_ITEMS), 66):
        for col in range(1, 8):
            style_input(ws.cell(row=extra_r, column=col))
        c = ws.cell(row=extra_r, column=8)
        c.value = f"=IF(D{extra_r}=\"\",\"\",D{extra_r}*E{extra_r})"
        c.number_format = "#,##0"
        style_formula(c)

    set_col_width(ws, {1: 10, 2: 30, 3: 14, 4: 10, 5: 10, 6: 10, 7: 12, 8: 14})
    set_row_height(ws, {1: 22, 2: 32})
    setup_sheet(ws, freeze_row=3, freeze_col=2)


def build_monthly_sheet(ws):
    ws.tab_color = COLORS["input"]

    style_title(ws, "月次徴収 ─ 口座引落結果の入力（充当額はマイナスで入力）",
                row=1, merge_cols=15, size=12)

    # ヘッダー行2: セクション
    round_sections = [(4, "1回目", 2), (6, "2回目", 2), (8, "3回目", 2),
                      (10, "4回目", 2), (12, "5回目", 2)]
    for col, label, span in round_sections:
        c = ws.cell(row=2, column=col)
        c.value = label
        c.fill = fill("header_mid")
        c.font = font(bold=True, size=9, color="white")
        c.alignment = align("center")
        c.border = border()
        ws.merge_cells(start_row=2, start_column=col,
                       end_row=2, end_column=col + span - 1)

    for col, label in [(1,""), (2,""), (3,""), (14,""), (15,"")]:
        c = ws.cell(row=2, column=col)
        c.fill = fill("header_mid")
        c.border = border()

    # ヘッダー行3
    headers = [
        "生徒番号", "氏名(自動)", "クラス(自動)",
        "請求", "充当",
        "請求", "充当",
        "請求", "充当",
        "請求", "充当",
        "請求", "充当",
        "現金振込", "年間合計\n(自動)",
    ]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), h,
                     color="header_light", font_color="black", size=9)

    # データ行
    for i, m in enumerate(MONTHLY, start=MT_DATA_START):
        sid = m[0]
        vals_input = [m[1], m[2], m[3], m[4], m[5], m[6],
                      m[7], m[8], m[9], m[10], m[11]]

        # 生徒番号
        c = ws.cell(row=i, column=1)
        c.value = sid
        c.number_format = "@"
        style_input(c)

        # 氏名 (XLOOKUP)
        c = ws.cell(row=i, column=2)
        c.value = f"=IFERROR(XLOOKUP(A{i},生徒マスター!$A:$A,生徒マスター!$B:$B),\"\")"
        style_formula(c)

        # クラス (XLOOKUP)
        c = ws.cell(row=i, column=3)
        c.value = f"=IFERROR(XLOOKUP(A{i},生徒マスター!$A:$A,生徒マスター!$C:$C),\"\")"
        style_formula(c)

        # 請求・充当 (4〜13列目)
        for j, v in enumerate(vals_input, start=4):
            c = ws.cell(row=i, column=j)
            c.value = v
            c.number_format = "#,##0"
            # 充当列（偶数列: 5,7,9,11,13）は水色
            if (j - 4) % 2 == 1:
                style_input(c)
            else:
                c.fill = fill("header_light")
                c.border = border()

        # 現金振込
        c = ws.cell(row=i, column=14)
        c.value = m[11]
        c.number_format = "#,##0"
        style_input(c)

        # 年間合計 (充当の合計 + 現金振込)
        c = ws.cell(row=i, column=15)
        c.value = f"=IFERROR(E{i}+G{i}+I{i}+K{i}+M{i}+N{i},0)"
        c.number_format = "#,##0"
        style_formula(c)

    # 空白テンプレート行
    for extra_r in range(MT_DATA_START + len(MONTHLY), MT_DATA_START + 301):
        ws.cell(row=extra_r, column=1).number_format = "@"
        style_input(ws.cell(row=extra_r, column=1))
        ws.cell(row=extra_r, column=2).value = (
            f"=IFERROR(XLOOKUP(A{extra_r},生徒マスター!$A:$A,生徒マスター!$B:$B),\"\")"
        )
        style_formula(ws.cell(row=extra_r, column=2))
        ws.cell(row=extra_r, column=3).value = (
            f"=IFERROR(XLOOKUP(A{extra_r},生徒マスター!$A:$A,生徒マスター!$C:$C),\"\")"
        )
        style_formula(ws.cell(row=extra_r, column=3))
        for col in [4, 6, 8, 10, 12]:
            ws.cell(row=extra_r, column=col).fill = fill("header_light")
            ws.cell(row=extra_r, column=col).border = border()
        for col in [5, 7, 9, 11, 13, 14]:
            style_input(ws.cell(row=extra_r, column=col))
        ws.cell(row=extra_r, column=15).value = (
            f"=IFERROR(E{extra_r}+G{extra_r}+I{extra_r}+K{extra_r}+M{extra_r}+N{extra_r},0)"
        )
        style_formula(ws.cell(row=extra_r, column=15))
        for col in [5, 7, 9, 11, 13, 14, 15]:
            ws.cell(row=extra_r, column=col).number_format = "#,##0"

    set_col_width(ws, {
        1: 10, 2: 14, 3: 8,
        4: 9, 5: 9, 6: 9, 7: 9, 8: 9, 9: 9,
        10: 9, 11: 9, 12: 9, 13: 9,
        14: 10, 15: 12,
    })
    set_row_height(ws, {1: 22, 2: 16, 3: 32})
    setup_sheet(ws, freeze_row=MT_DATA_START, freeze_col=3)


def build_expense_record_sheet(ws):
    ws.tab_color = COLORS["red"]
    style_title(ws, "支出記録 ─ 支出が発生するたびに1行追加（D列空欄=全員、番号入力=特定生徒）",
                row=1, merge_cols=6, size=12)

    col_headers = [
        "日付", "支出番号", "項目名(自動)", "生徒番号\n(空=全員)", "金額(円)", "備考",
    ]
    for col, h in enumerate(col_headers, start=1):
        style_header(ws.cell(row=2, column=col), h, color="red",
                     font_color="white", size=10)

    for i, rec in enumerate(EXPENSE_RECORDS, start=3):
        date_str, item_num, student_id, amount, note = rec
        # 日付
        c = ws.cell(row=i, column=1)
        c.value = date_str
        c.number_format = "yyyy/mm/dd"
        style_input(c)
        # 支出番号
        c = ws.cell(row=i, column=2)
        c.value = item_num
        style_input(c)
        # 項目名 (XLOOKUP)
        c = ws.cell(row=i, column=3)
        c.value = f"=IFERROR(XLOOKUP(B{i},支出項目マスター!$A:$A,支出項目マスター!$B:$B),\"\")"
        style_formula(c)
        # 生徒番号
        c = ws.cell(row=i, column=4)
        c.value = student_id
        c.number_format = "@"
        style_input(c)
        # 金額
        c = ws.cell(row=i, column=5)
        c.value = amount
        c.number_format = "#,##0"
        style_input(c)
        # 備考
        c = ws.cell(row=i, column=6)
        c.value = note
        style_input(c)

    # 空白テンプレート行（500行分）
    for extra_r in range(3 + len(EXPENSE_RECORDS), 504):
        ws.cell(extra_r, 1).number_format = "yyyy/mm/dd"
        style_input(ws.cell(extra_r, 1))
        style_input(ws.cell(extra_r, 2))
        ws.cell(extra_r, 3).value = (
            f"=IFERROR(XLOOKUP(B{extra_r},支出項目マスター!$A:$A,支出項目マスター!$B:$B),\"\")"
        )
        style_formula(ws.cell(extra_r, 3))
        ws.cell(extra_r, 4).number_format = "@"
        style_input(ws.cell(extra_r, 4))
        ws.cell(extra_r, 5).number_format = "#,##0"
        style_input(ws.cell(extra_r, 5))
        style_input(ws.cell(extra_r, 6))

    set_col_width(ws, {1: 13, 2: 10, 3: 28, 4: 12, 5: 12, 6: 30})
    set_row_height(ws, {1: 22, 2: 32})
    setup_sheet(ws, freeze_row=3, freeze_col=3)


def _expense_formula(item_num, row):
    """支出項目ごとの金額計算数式（給付型判定付き）"""
    return (
        f'=IF(AND('
        f'IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$K:$K,""),"")&""<>"",'
        f'IFERROR(XLOOKUP({item_num},支出項目マスター!$A:$A,支出項目マスター!$G:$G,""),"")&""<>""'
        f'),"給",'
        f'IFERROR('
        f'IF(SUMIFS(支出記録!$E:$E,支出記録!$B:$B,{item_num},支出記録!$D:$D,$A{row})>0,'
        f'SUMIFS(支出記録!$E:$E,支出記録!$B:$B,{item_num},支出記録!$D:$D,$A{row}),'
        f'SUMIFS(支出記録!$E:$E,支出記録!$B:$B,{item_num},支出記録!$D:$D,"")),'
        f'0))'
    )


def build_individual_output_sheet(ws):
    ws.tab_color = COLORS["purple"]
    style_title(ws, "個人別管理表(出力) ─ 自動生成 ※このシートは触らないでください",
                row=1, merge_cols=20, size=12)

    # セクションヘッダー（行2）
    sections = [
        (1,  "基本情報",     5),
        (6,  "繰越",         1),
        (7,  "月次徴収",     2),
        (9,  "支出の部",    len(EXPENSE_ITEMS)),
        (9 + len(EXPENSE_ITEMS), "集計", 3),
    ]
    for start_col, label, span in sections:
        c = ws.cell(row=2, column=start_col)
        c.value = label
        c.fill = fill("header_mid")
        c.font = font(bold=True, size=9, color="white")
        c.border = border()
        c.alignment = align("center")
        if span > 1:
            ws.merge_cells(start_row=2, start_column=start_col,
                           end_row=2, end_column=start_col + span - 1)

    # 列ヘッダー（行3）
    base_headers = [
        "生徒番号", "氏名", "クラス", "ふりがな", "給付型",
        "前年度繰越金",
        "月次徴収\n年間合計", "現金振込",
    ]
    for col, h in enumerate(base_headers, start=1):
        style_header(ws.cell(row=3, column=col), h,
                     color="header_light", font_color="black", size=9, wrap=True)

    # 支出項目ヘッダー
    EXPENSE_COL_START = 9
    for idx, item in enumerate(EXPENSE_ITEMS):
        col = EXPENSE_COL_START + idx
        c = ws.cell(row=3, column=col)
        c.value = f"[{item[0]}]\n{item[1]}"
        c.fill = fill("header_light")
        c.font = font(bold=False, size=8)
        c.border = border()
        c.alignment = align("center", "center", wrap=True)

    # 集計ヘッダー
    TOTAL_COL = EXPENSE_COL_START + len(EXPENSE_ITEMS)
    for offset, label in enumerate(["総収入(A)", "支出総額(B)", "次年度繰越金\n(A-B)"]):
        c = ws.cell(row=3, column=TOTAL_COL + offset)
        style_header(c, label, color="header", font_color="white", size=9)

    # データ行
    DATA_START = 4
    for i, s in enumerate(STUDENTS, start=DATA_START):
        row = i
        sid = s[0]

        # A: 生徒番号
        c = ws.cell(row=row, column=1)
        c.value = sid
        c.number_format = "@"
        c.fill = fill("gray")
        c.border = border()
        c.alignment = align("center")

        # B: 氏名
        c = ws.cell(row=row, column=2)
        c.value = f"=IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$B:$B),\"\")"
        style_formula(c)

        # C: クラス
        c = ws.cell(row=row, column=3)
        c.value = f"=IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$C:$C),\"\")"
        style_formula(c)
        c.alignment = align("center")

        # D: ふりがな
        c = ws.cell(row=row, column=4)
        c.value = f"=IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$D:$D),\"\")"
        style_formula(c)

        # E: 給付型フラグ
        c = ws.cell(row=row, column=5)
        c.value = f"=IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$K:$K),\"\")"
        style_formula(c)
        c.alignment = align("center")

        # F: 前年度繰越金
        c = ws.cell(row=row, column=6)
        c.value = f"=IFERROR(XLOOKUP($A{row},生徒マスター!$A:$A,生徒マスター!$M:$M),0)"
        c.number_format = "#,##0"
        style_formula(c)

        # G: 月次徴収年間合計
        c = ws.cell(row=row, column=7)
        c.value = (
            f"=IFERROR(XLOOKUP($A{row},'月次徴収'!$A:$A,'月次徴収'!$O:$O),0)"
        )
        c.number_format = "#,##0"
        style_formula(c)

        # H: 現金振込
        c = ws.cell(row=row, column=8)
        c.value = (
            f"=IFERROR(XLOOKUP($A{row},'月次徴収'!$A:$A,'月次徴収'!$N:$N),0)"
        )
        c.number_format = "#,##0"
        style_formula(c)

        # 支出項目列
        for idx, item in enumerate(EXPENSE_ITEMS):
            col = EXPENSE_COL_START + idx
            c = ws.cell(row=row, column=col)
            c.value = _expense_formula(item[0], row)
            c.number_format = '#,##0;-#,##0;""'
            style_formula(c)

        # 総収入(A) = 前年度繰越 + 月次合計
        c = ws.cell(row=row, column=TOTAL_COL)
        c.value = f"=F{row}+G{row}"
        c.number_format = "#,##0"
        c.fill = fill("green_light")
        c.border = border()
        c.font = font(bold=True, size=10)
        c.alignment = align("right")

        # 支出総額(B) = SUM of expense cols (文字列「給」はSUMで0扱い)
        exp_start_col = get_column_letter(EXPENSE_COL_START)
        exp_end_col = get_column_letter(EXPENSE_COL_START + len(EXPENSE_ITEMS) - 1)
        c = ws.cell(row=row, column=TOTAL_COL + 1)
        c.value = f"=SUMPRODUCT(IF(ISNUMBER({exp_start_col}{row}:{exp_end_col}{row}),{exp_start_col}{row}:{exp_end_col}{row},0))"
        c.number_format = "#,##0"
        c.fill = fill("orange_light")
        c.border = border()
        c.font = font(bold=True, size=10)
        c.alignment = align("right")

        # 次年度繰越金(A-B)
        c = ws.cell(row=row, column=TOTAL_COL + 2)
        c.value = f"={get_column_letter(TOTAL_COL)}{row}-{get_column_letter(TOTAL_COL+1)}{row}"
        c.number_format = "#,##0"
        c.fill = fill("yellow_light")
        c.border = border()
        c.font = font(bold=True, size=10)
        c.alignment = align("right")

    # 空白テンプレート行 (250名分)
    EMPTY_START = DATA_START + len(STUDENTS)
    for extra_r in range(EMPTY_START, EMPTY_START + 250):
        ws.cell(row=extra_r, column=1).number_format = "@"
        ws.cell(row=extra_r, column=1).fill = fill("gray")
        ws.cell(row=extra_r, column=1).border = border()
        ws.cell(row=extra_r, column=1).alignment = align("center")

        for col in range(2, TOTAL_COL + 3):
            c = ws.cell(row=extra_r, column=col)
            style_formula(c)

        ws.cell(row=extra_r, column=2).value = (
            f'=IFERROR(XLOOKUP($A{extra_r},生徒マスター!$A:$A,生徒マスター!$B:$B),"")')
        ws.cell(row=extra_r, column=3).value = (
            f'=IFERROR(XLOOKUP($A{extra_r},生徒マスター!$A:$A,生徒マスター!$C:$C),"")')
        ws.cell(row=extra_r, column=4).value = (
            f'=IFERROR(XLOOKUP($A{extra_r},生徒マスター!$A:$A,生徒マスター!$D:$D),"")')
        ws.cell(row=extra_r, column=5).value = (
            f'=IFERROR(XLOOKUP($A{extra_r},生徒マスター!$A:$A,生徒マスター!$K:$K),"")')
        ws.cell(row=extra_r, column=6).value = (
            f'=IFERROR(XLOOKUP($A{extra_r},生徒マスター!$A:$A,生徒マスター!$M:$M),0)')
        ws.cell(row=extra_r, column=6).number_format = "#,##0"
        ws.cell(row=extra_r, column=7).value = (
            f"=IFERROR(XLOOKUP($A{extra_r},'月次徴収'!$A:$A,'月次徴収'!$O:$O),0)")
        ws.cell(row=extra_r, column=7).number_format = "#,##0"
        ws.cell(row=extra_r, column=8).value = (
            f"=IFERROR(XLOOKUP($A{extra_r},'月次徴収'!$A:$A,'月次徴収'!$N:$N),0)")
        ws.cell(row=extra_r, column=8).number_format = "#,##0"

        for idx, item in enumerate(EXPENSE_ITEMS):
            col = EXPENSE_COL_START + idx
            c = ws.cell(row=extra_r, column=col)
            c.value = _expense_formula(item[0], extra_r)
            c.number_format = '#,##0;-#,##0;""'

        ws.cell(row=extra_r, column=TOTAL_COL).value = f"=F{extra_r}+G{extra_r}"
        ws.cell(row=extra_r, column=TOTAL_COL).number_format = "#,##0"
        ws.cell(row=extra_r, column=TOTAL_COL).fill = fill("green_light")
        ws.cell(row=extra_r, column=TOTAL_COL).border = border()
        ws.cell(row=extra_r, column=TOTAL_COL).font = font(bold=True)

        ws.cell(row=extra_r, column=TOTAL_COL + 1).value = (
            f"=SUMPRODUCT(IF(ISNUMBER({exp_start_col}{extra_r}:{exp_end_col}{extra_r}),"
            f"{exp_start_col}{extra_r}:{exp_end_col}{extra_r},0))"
        )
        ws.cell(row=extra_r, column=TOTAL_COL + 1).number_format = "#,##0"
        ws.cell(row=extra_r, column=TOTAL_COL + 1).fill = fill("orange_light")
        ws.cell(row=extra_r, column=TOTAL_COL + 1).border = border()
        ws.cell(row=extra_r, column=TOTAL_COL + 1).font = font(bold=True)

        ws.cell(row=extra_r, column=TOTAL_COL + 2).value = (
            f"={get_column_letter(TOTAL_COL)}{extra_r}-{get_column_letter(TOTAL_COL+1)}{extra_r}"
        )
        ws.cell(row=extra_r, column=TOTAL_COL + 2).number_format = "#,##0"
        ws.cell(row=extra_r, column=TOTAL_COL + 2).fill = fill("yellow_light")
        ws.cell(row=extra_r, column=TOTAL_COL + 2).border = border()
        ws.cell(row=extra_r, column=TOTAL_COL + 2).font = font(bold=True)

    # 列幅設定
    col_widths = {1: 10, 2: 14, 3: 7, 4: 14, 5: 7, 6: 12, 7: 12, 8: 10}
    for idx in range(len(EXPENSE_ITEMS)):
        col_widths[EXPENSE_COL_START + idx] = 10
    col_widths[TOTAL_COL]     = 13
    col_widths[TOTAL_COL + 1] = 13
    col_widths[TOTAL_COL + 2] = 14
    set_col_width(ws, col_widths)
    set_row_height(ws, {1: 22, 2: 16, 3: 40})
    setup_sheet(ws, freeze_row=DATA_START, freeze_col=4)


def build_settlement_sheet(ws):
    ws.tab_color = COLORS["red"]

    # タイトル
    style_title(ws, "精 算 書", row=1, merge_cols=4, size=16)

    # 入力欄
    ws.cell(row=3, column=1).value = "生徒番号を入力 →"
    ws.cell(row=3, column=1).font = font(bold=True, size=11)
    ws.cell(row=3, column=1).alignment = align("right", "center")

    c = ws.cell(row=3, column=2)
    c.value = ""  # ← ここに生徒番号を入力
    c.fill = fill("yellow")
    c.border = border("medium")
    c.font = font(bold=True, size=12)
    c.alignment = align("center", "center")
    c.number_format = "@"

    ws.cell(row=3, column=3).value = "← 黄色のセルに生徒番号(5桁)を入力してください"
    ws.cell(row=3, column=3).font = font(size=9, color="red")
    ws.merge_cells("C3:D3")

    # 生徒情報の自動表示
    info_rows = [
        (5,  "生徒番号",   "=$B$3"),
        (6,  "氏　　名",   "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$B:$B),\"\")"),
        (7,  "クラス",     "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$C:$C),\"\")"),
        (8,  "保護者名",   "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$E:$E),\"\")"),
        (9,  "在籍状態",   "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$L:$L),\"\")"),
        (10, "給付型",     "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$K:$K),\"\")"),
    ]
    for row, label, formula in info_rows:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = fill("header_light")
        ws.cell(row=row, column=1).border = border()
        ws.cell(row=row, column=1).alignment = align("right", "center")
        c = ws.cell(row=row, column=2)
        c.value = formula
        c.font = font(size=11)
        c.fill = fill("gray")
        c.border = border()
        c.alignment = align("left", "center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    # 収入の部ヘッダー
    ws.cell(row=12, column=1).value = "【収入の部】"
    ws.cell(row=12, column=1).font = font(bold=True, size=11, color="header")
    ws.cell(row=12, column=1).alignment = align("left")
    ws.merge_cells("A12:D12")

    income_items = [
        (13, "前年度繰越金",
         "=IFERROR(XLOOKUP($B$3,生徒マスター!$A:$A,生徒マスター!$M:$M),0)"),
        (14, "月次徴収 年間合計",
         "=IFERROR(XLOOKUP($B$3,'月次徴収'!$A:$A,'月次徴収'!$O:$O),0)"),
        (15, "現金振込",
         "=IFERROR(XLOOKUP($B$3,'月次徴収'!$A:$A,'月次徴収'!$N:$N),0)"),
    ]
    for row, label, formula in income_items:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = font(size=10)
        ws.cell(row=row, column=1).border = border()
        ws.cell(row=row, column=1).fill = fill("gray")
        ws.cell(row=row, column=1).alignment = align("right", "center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=4)
        c.value = formula
        c.number_format = "#,##0"
        c.fill = fill("gray")
        c.border = border()
        c.alignment = align("right", "center")

    ws.cell(row=16, column=1).value = "収入合計"
    ws.cell(row=16, column=1).font = font(bold=True, size=11)
    ws.cell(row=16, column=1).fill = fill("green_light")
    ws.cell(row=16, column=1).border = border("medium")
    ws.merge_cells("A16:C16")
    c = ws.cell(row=16, column=4)
    c.value = "=D13+D14+D15"
    c.number_format = "#,##0"
    c.fill = fill("green_light")
    c.border = border("medium")
    c.font = font(bold=True, size=12)
    c.alignment = align("right", "center")

    # 支出の部ヘッダー
    ws.cell(row=18, column=1).value = "【支出の部】"
    ws.cell(row=18, column=1).font = font(bold=True, size=11, color="header")
    ws.merge_cells("A18:D18")

    # 支出項目（行19〜 各項目ごとに1行）
    EXPENSE_DETAIL_START = 19
    for idx, item in enumerate(EXPENSE_ITEMS):
        row = EXPENSE_DETAIL_START + idx
        item_num, item_name = item[0], item[1]
        ws.cell(row=row, column=1).value = item_name
        ws.cell(row=row, column=1).font = font(size=10)
        ws.cell(row=row, column=1).border = border()
        ws.cell(row=row, column=1).alignment = align("left", "center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        c = ws.cell(row=row, column=4)
        c.value = _expense_formula(item_num, row).replace(f"$A{row}", "$B$3")
        c.number_format = '#,##0;-#,##0;""'
        c.fill = fill("gray")
        c.border = border()
        c.alignment = align("right", "center")

    TOTAL_ROW = EXPENSE_DETAIL_START + len(EXPENSE_ITEMS)
    exp_range = f"D{EXPENSE_DETAIL_START}:D{TOTAL_ROW - 1}"
    ws.cell(row=TOTAL_ROW, column=1).value = "支出合計"
    ws.cell(row=TOTAL_ROW, column=1).font = font(bold=True, size=11)
    ws.cell(row=TOTAL_ROW, column=1).fill = fill("orange_light")
    ws.cell(row=TOTAL_ROW, column=1).border = border("medium")
    ws.merge_cells(f"A{TOTAL_ROW}:C{TOTAL_ROW}")
    c = ws.cell(row=TOTAL_ROW, column=4)
    c.value = f"=SUMPRODUCT(IF(ISNUMBER({exp_range}),{exp_range},0))"
    c.number_format = "#,##0"
    c.fill = fill("orange_light")
    c.border = border("medium")
    c.font = font(bold=True, size=12)
    c.alignment = align("right", "center")

    # 精算金額
    BALANCE_ROW = TOTAL_ROW + 2
    ws.cell(row=BALANCE_ROW, column=1).value = "【精算金額（次年度繰越）】"
    ws.cell(row=BALANCE_ROW, column=1).font = font(bold=True, size=12, color="header")
    ws.merge_cells(f"A{BALANCE_ROW}:C{BALANCE_ROW}")
    c = ws.cell(row=BALANCE_ROW, column=4)
    c.value = f"=D16-D{TOTAL_ROW}"
    c.number_format = "#,##0"
    c.fill = fill("yellow")
    c.border = border("medium")
    c.font = font(bold=True, size=13)
    c.alignment = align("right", "center")

    set_col_width(ws, {1: 28, 2: 12, 3: 12, 4: 14})
    setup_sheet(ws, freeze_row=5, freeze_col=1, zoom=100)


def build_financial_sheet(ws):
    ws.tab_color = COLORS["navy"]
    style_title(ws, "決 算 書（出力）─ 自動集計 ※このシートは触らないでください",
                row=1, merge_cols=5, size=14)

    # ヘッダー
    headers = ["番号", "支出項目名", "教科・区分", "単価", "対象人数", "支出合計（自動）"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=2, column=col), h, color="navy", font_color="white")

    for idx, item in enumerate(EXPENSE_ITEMS, start=3):
        num, name, category, price, count, timing, kyufu = item
        ws.cell(row=idx, column=1).value = num
        ws.cell(row=idx, column=1).border = border()
        ws.cell(row=idx, column=1).alignment = align("center")

        ws.cell(row=idx, column=2).value = name
        ws.cell(row=idx, column=2).border = border()

        ws.cell(row=idx, column=3).value = category
        ws.cell(row=idx, column=3).border = border()
        ws.cell(row=idx, column=3).alignment = align("center")

        ws.cell(row=idx, column=4).value = price
        ws.cell(row=idx, column=4).number_format = "#,##0"
        ws.cell(row=idx, column=4).border = border()
        ws.cell(row=idx, column=4).alignment = align("right")

        ws.cell(row=idx, column=5).value = count
        ws.cell(row=idx, column=5).border = border()
        ws.cell(row=idx, column=5).alignment = align("center")

        # 支出合計 = SUMIF on 支出記録 for this item number
        c = ws.cell(row=idx, column=6)
        c.value = (
            f"=SUMPRODUCT(IF(支出記録!$B$3:$B$503={num},"
            f"IF(ISNUMBER(支出記録!$E$3:$E$503),支出記録!$E$3:$E$503,0),0))"
        )
        c.number_format = "#,##0"
        c.fill = fill("gray")
        c.border = border()
        c.alignment = align("right")

    GRAND_TOTAL_ROW = 3 + len(EXPENSE_ITEMS)
    ws.cell(row=GRAND_TOTAL_ROW, column=1).value = "合　計"
    ws.cell(row=GRAND_TOTAL_ROW, column=1).font = font(bold=True, size=12)
    ws.cell(row=GRAND_TOTAL_ROW, column=1).fill = fill("header_light")
    ws.cell(row=GRAND_TOTAL_ROW, column=1).border = border("medium")
    ws.merge_cells(f"A{GRAND_TOTAL_ROW}:E{GRAND_TOTAL_ROW}")

    c = ws.cell(row=GRAND_TOTAL_ROW, column=6)
    c.value = f"=SUM(F3:F{GRAND_TOTAL_ROW-1})"
    c.number_format = "#,##0"
    c.fill = fill("header_light")
    c.border = border("medium")
    c.font = font(bold=True, size=12)
    c.alignment = align("right")

    set_col_width(ws, {1: 8, 2: 30, 3: 14, 4: 10, 5: 10, 6: 16})
    set_row_height(ws, {1: 26, 2: 22})
    setup_sheet(ws, freeze_row=3, freeze_col=2)


def create_workbook(output_path: str):
    wb = Workbook()

    ws_guide = wb.active
    ws_guide.title = "操作ガイド"
    build_guide_sheet(ws_guide)

    ws_sm = wb.create_sheet("生徒マスター")
    build_student_master_sheet(ws_sm)

    ws_em = wb.create_sheet("支出項目マスター")
    build_expense_master_sheet(ws_em)

    ws_mt = wb.create_sheet("月次徴収")
    build_monthly_sheet(ws_mt)

    ws_er = wb.create_sheet("支出記録")
    build_expense_record_sheet(ws_er)

    ws_ind = wb.create_sheet("個人別管理表(出力)")
    build_individual_output_sheet(ws_ind)

    ws_set = wb.create_sheet("精算書(出力)")
    build_settlement_sheet(ws_set)

    ws_fin = wb.create_sheet("決算書(出力)")
    build_financial_sheet(ws_fin)

    wb.save(output_path)
    print(f"  ✓ master_マスターブック.xlsx → {output_path}")


if __name__ == "__main__":
    os.makedirs("../output", exist_ok=True)
    create_workbook("../output/master_マスターブック.xlsx")
