# -*- coding: utf-8 -*-
"""ツール① 番号紐付けテンプレート生成"""

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys, os

sys.path.insert(0, os.path.dirname(__file__))
from styles import (fill, font, border, align, style_header, style_input,
                    style_formula, style_title, style_guide_row,
                    setup_sheet, set_col_width, set_row_height, COLORS)

SAMPLE_STUDENTS = [
    ("04001","山田 太郎", "やまだ たろう",   "R4","Ａ",1,"Ｂ",3,"Ｃ",5),
    ("04002","鈴木 花子", "すずき はなこ",   "R4","Ａ",2,"Ａ",1,"Ａ",2),
    ("04003","田中 一郎", "たなか いちろう", "R4","Ｂ",1,"Ｃ",4,"Ｂ",7),
    ("04004","佐藤 美咲", "さとう みさき",   "R4","Ｂ",2,"Ｄ",2,"Ｄ",1),
    ("04005","高橋 健一", "たかはし けんいち","R4","Ｃ",1,"Ｅ",5,"Ｅ",3),
    ("05001","伊藤 奈々", "いとう なな",     "R5","Ａ",1,"Ｂ",2,"",  ""),
    ("05002","中村 大輔", "なかむら だいすけ","R5","Ｂ",3,"Ａ",4,"",  ""),
    ("05003","小林 あい", "こばやし あい",   "R5","Ｃ",2,"Ｃ",1,"",  ""),
    ("06001","加藤 翼",   "かとう つばさ",   "R6","Ａ",1,"",  "","",  ""),
    ("06002","吉田 愛",   "よしだ あい",     "R6","Ｂ",2,"",  "","",  ""),
]

SAMPLE_NEW_ROSTER = [
    ("山田 太郎","Ｄ",2),
    ("鈴木 花子","Ａ",1),
    ("田中 一郎","Ｆ",4),
    ("佐藤 美咲","Ｂ",8),
    ("高橋 健一","Ｃ",3),
]


def build_guide_sheet(ws):
    ws.tab_color = COLORS["header"]
    style_title(ws, "操作ガイド ─ 番号紐付けテンプレート", row=1, merge_cols=4, size=14)
    rows = [
        ("【目的】",
         "クラス替えが確定したとき、新クラスと出席番号を生徒マスターへ\n"
         "正確に転記するためのツールです。名前の一致/不一致を自動判定します。"),
        ("【年度更新手順】",
         "Step1: 先生から新クラス名簿を受領\n"
         "Step2: 「新年度名簿入力」シート（橙）にA〜C列を貼り付け\n"
         "Step3: 「照合結果」シート（赤）で全員「一致」を確認\n"
         "Step4: 確認後、照合結果B/C列を生徒マスターへ値貼り付け"),
        ("【照合状態の見方】",
         "一致（緑）     → そのまま転記可能\n"
         "氏名不一致（黄）→ ふりがな・旧クラスで確認\n"
         "未マッチ（赤） → 転入生なら手動追加、表記ゆれなら氏名を修正"),
        ("⚠ 注意",
         "「山田 太郎」と「山田太郎」（スペースなし）は別人と判定されます。\n"
         "名前の表記を個人別管理表と揃えてください。"),
    ]
    for i, (label, content) in enumerate(rows, start=3):
        style_guide_row(ws, i, label, content)
        ws.row_dimensions[i].height = 52
    set_col_width(ws, {"A": 22, "B": 70})
    setup_sheet(ws, freeze_row=3)


def build_master_sheet(ws):
    ws.tab_color = COLORS["green"]
    style_title(ws, "生徒マスター ─ 年度別クラス・出席番号管理（主キー: 生徒番号）",
                row=1, merge_cols=14, size=12)

    sections = [(1,"基本情報",5),(6,"1年次",2),(8,"2年次",2),(10,"3年次",2),(12,"次年度用",2),(14,"備考",1)]
    for sc, lbl, sp in sections:
        c = ws.cell(row=2, column=sc)
        c.value = lbl
        c.fill = fill("header_mid")
        c.font = font(bold=True, size=9, color="white")
        c.alignment = align("center")
        c.border = border()
        if sp > 1:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc+sp-1)

    for col, h in enumerate(
        ["No.","生徒番号","氏名","ふりがな","入学年度","クラス","番号",
         "クラス","番号","クラス","番号","新クラス","新番号","備考"], start=1):
        style_header(ws.cell(row=3, column=col), h, color="header_light",
                     font_color="black", bold=True, size=9)

    for i, s in enumerate(SAMPLE_STUDENTS, start=4):
        ws.cell(row=i, column=1).value = i - 3
        ws.cell(row=i, column=1).alignment = align("center")
        for col, val in enumerate(
            [s[0],s[1],s[2],s[3],s[4],s[5],s[6],s[7],s[8],s[9],"","","",""], start=2):
            c = ws.cell(row=i, column=col)
            c.value = val if val != "" else None
            if col in (2,):
                c.number_format = "@"
            color = "yellow_light" if col in (12,13,14) else "input"
            style_input(c, color=color)

    for r in range(4 + len(SAMPLE_STUDENTS), 305):
        ws.cell(row=r, column=1).value = r - 3
        ws.cell(row=r, column=1).alignment = align("center")
        for col in range(2, 15):
            c = ws.cell(row=r, column=col)
            if col == 2:
                c.number_format = "@"
            style_input(c, color="yellow_light" if col in (12,13,14) else "input")

    dv = DataValidation(type="list", formula1='"Ａ,Ｂ,Ｃ,Ｄ,Ｅ,Ｆ"', showDropDown=False)
    for col in [6,8,10,12]:
        dv2 = DataValidation(type="list", formula1='"Ａ,Ｂ,Ｃ,Ｄ,Ｅ,Ｆ"', showDropDown=False)
        cl = get_column_letter(col)
        dv2.sqref = f"{cl}4:{cl}304"
        ws.add_data_validation(dv2)

    set_col_width(ws, {1:5,2:10,3:16,4:18,5:10,6:8,7:7,8:8,9:7,10:8,11:7,12:10,13:8,14:20})
    set_row_height(ws, {1:22,2:16,3:20})
    setup_sheet(ws, freeze_row=4, freeze_col=3)


def build_new_roster_sheet(ws):
    ws.tab_color = COLORS["orange"]
    style_title(ws, "新年度名簿入力 ─ 先生からもらった名簿をここに貼り付け",
                row=1, merge_cols=5, size=13)

    note = ("★ 貼り付け方法: 名簿をコピーして A列から「右クリック→値として貼り付け」\n"
            "★ 注意: 氏名は個人別管理表と同じ書き方（スペース有無・全半角）に揃えてください")
    ws.cell(row=2, column=1).value = note
    ws.cell(row=2, column=1).font = font(size=9, color="red")
    ws.cell(row=2, column=1).alignment = align("left","center",wrap=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 36

    for col, h in enumerate(["氏名","新クラス","新出席番号","（旧クラス参考）","（ふりがな参考）"], start=1):
        style_header(ws.cell(row=3, column=col), h, color="orange", font_color="white", size=10)

    for i, row_data in enumerate(SAMPLE_NEW_ROSTER, start=4):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col)
            c.value = val
            style_input(c)

    for r in range(4 + len(SAMPLE_NEW_ROSTER), 304):
        for col in range(1, 4):
            style_input(ws.cell(row=r, column=col))

    set_col_width(ws, {1:18,2:10,3:12,4:14,5:18})
    set_row_height(ws, {1:22,2:36,3:20})
    setup_sheet(ws, freeze_row=4)


def build_match_result_sheet(ws):
    ws.tab_color = COLORS["red"]
    style_title(ws, "照合結果 ─ 自動照合で確認 → 生徒マスターへ転記",
                row=1, merge_cols=7, size=13)

    legend = [("一致（緑）","氏名が正確に一致。そのまま転記可能"),
              ("氏名不一致（黄）","似ているが別人の可能性。ふりがな・旧クラスで確認"),
              ("未マッチ（赤）","マスターに存在しない。転入生 or 表記ゆれ")]
    for i, (k, v) in enumerate(legend, start=2):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=1).font = font(bold=True, size=9)
        ws.cell(row=i, column=2).value = v
        ws.cell(row=i, column=2).font = font(size=9)
        ws.row_dimensions[i].height = 16
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

    for col, h in enumerate(
        ["氏名（名簿から）","新クラス","新出席番号","生徒番号（自動）","照合状態","旧クラス（参考）","ふりがな（参考）"],
        start=1):
        style_header(ws.cell(row=5, column=col), h, color="header", size=9)

    MASTER = "生徒マスター"
    ROSTER = "新年度名簿入力"
    MAX = 303

    for r in range(6, MAX + 6):
        rr = r - 5 + 3  # 名簿入力シートの対応行

        ws.cell(row=r, column=1).value = f"=IF('{ROSTER}'!A{rr}=\"\",\"\",'{ROSTER}'!A{rr})"
        ws.cell(row=r, column=2).value = f"=IF('{ROSTER}'!B{rr}=\"\",\"\",'{ROSTER}'!B{rr})"
        ws.cell(row=r, column=3).value = f"=IF('{ROSTER}'!C{rr}=\"\",\"\",'{ROSTER}'!C{rr})"
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = border()
            ws.cell(row=r, column=col).alignment = align("center" if col > 1 else "left", "center")

        # 生徒番号 (XLOOKUP: 氏名→生徒番号)
        c = ws.cell(row=r, column=4)
        c.value = (f'=IF(A{r}="","",IFERROR(XLOOKUP(A{r},'
                   f"'{MASTER}'!$C:$C,'{MASTER}'!$B:$B,\"未マッチ\"),\"未マッチ\"))")
        c.fill = fill("gray"); c.border = border()
        c.number_format = "@"; c.alignment = align("center","center")

        # 照合状態
        c = ws.cell(row=r, column=5)
        c.value = (f'=IF(A{r}="","",IF(D{r}="未マッチ","未マッチ",'
                   f'IF(IFERROR(XLOOKUP(D{r},\'{MASTER}\'!$B:$B,\'{MASTER}\'!$C:$C,""),"")'
                   f'=A{r},"一致","氏名不一致")))')
        c.fill = fill("gray"); c.border = border()
        c.font = font(bold=True); c.alignment = align("center","center")

        # 旧クラス参考
        c = ws.cell(row=r, column=6)
        c.value = (f'=IF(D{r}="未マッチ","",IFERROR(XLOOKUP(D{r},'
                   f"'{MASTER}'!$B:$B,'{MASTER}'!$C:$C,\"\"),\"\"))")
        c.fill = fill("gray"); c.border = border(); c.alignment = align("center","center")

        # ふりがな参考
        c = ws.cell(row=r, column=7)
        c.value = (f'=IF(D{r}="未マッチ","",IFERROR(XLOOKUP(D{r},'
                   f"'{MASTER}'!$B:$B,'{MASTER}'!$D:$D,\"\"),\"\"))")
        c.fill = fill("gray"); c.border = border(); c.alignment = align("left","center")

    data_range = f"A6:G{MAX+5}"
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f'$E6="一致"'],
                    fill=fill("green_light")))
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f'$E6="氏名不一致"'],
                    fill=fill("yellow_light")))
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f'$E6="未マッチ"'],
                    fill=fill("red_light")))

    set_col_width(ws, {1:18,2:10,3:12,4:12,5:14,6:12,7:18})
    set_row_height(ws, {1:22,5:20})
    setup_sheet(ws, freeze_row=6, freeze_col=4)


def create_workbook(output_path):
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "操作ガイド"
    build_guide_sheet(ws_guide)
    build_master_sheet(wb.create_sheet("生徒マスター"))
    build_new_roster_sheet(wb.create_sheet("新年度名簿入力"))
    build_match_result_sheet(wb.create_sheet("照合結果"))
    wb.save(output_path)
    print(f"  ✓ 番号紐付けテンプレート.xlsx → {output_path}")


if __name__ == "__main__":
    os.makedirs("../output", exist_ok=True)
    create_workbook("../output/番号紐付けテンプレート.xlsx")
