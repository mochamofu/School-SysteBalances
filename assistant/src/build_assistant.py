# -*- coding: utf-8 -*-
"""積立金入力アシスタント.xlsx を生成するスクリプト

このブックはUI（入力フォーム・結果表示）だけを持ち、実際の処理は
vba/ フォルダの .bas モジュール（インポートして使う）が行う。
セル番地は VBA 側の定数と一致させてあるので、レイアウトを変える場合は
必ず両方を揃えること。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "203864"
BLUE = "2F5FA8"
LBLUE = "DCE8FA"
YELLOW = "FFF3B0"
GREEN = "E2F6E6"
GRAY = "808080"

F_TITLE = Font(name="游ゴシック", size=16, bold=True, color=NAVY)
F_H = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
F_LABEL = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
F_BODY = Font(name="游ゴシック", size=11)
F_NOTE = Font(name="游ゴシック", size=10, color=GRAY)

FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_LBL = PatternFill("solid", fgColor=LBLUE)
FILL_IN = PatternFill("solid", fgColor=YELLOW)   # 入力セル＝黄色
FILL_OUT = PatternFill("solid", fgColor=GREEN)   # 自動出力セル＝緑

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header(ws, text, width=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F_TITLE
    ws.row_dimensions[1].height = 28


def label_input(ws, row, label, addr_note=None, input_col=3, note=None):
    """B列=ラベル、C列=黄色い入力セル の1行を作る"""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = F_LABEL
    lc.fill = FILL_LBL
    lc.border = BORDER
    ic = ws.cell(row=row, column=input_col)
    ic.fill = FILL_IN
    ic.border = BORDER
    ic.font = F_BODY
    if note:
        nc = ws.cell(row=row, column=input_col + 1, value=note)
        nc.font = F_NOTE
    return ic


def table_header(ws, row, cols, start_col=1):
    for i, h in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F_H
        c.fill = FILL_H
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")


def build():
    wb = Workbook()

    # ============ メニュー ============
    ws = wb.active
    ws.title = "メニュー"
    header(ws, "積立金会計 入力アシスタント", 10)
    ws["A2"] = "はじめて使うとき：VBAモジュール（A00〜A11.bas）をインポート後、Alt+F8→「初期設定」を1回実行するとボタンが並びます。"
    ws["A2"].font = F_NOTE
    ws["B3"] = "◆ 名簿・クラス替え"; ws["B3"].font = F_LABEL
    ws["E3"] = "◆ 日々の入力"; ws["E3"].font = F_LABEL
    ws["H3"] = "◆ 年度末・点検"; ws["H3"].font = F_LABEL
    ws["K3"] = "◆ 支出項目・引き継ぎ"; ws["K3"].font = F_LABEL
    for i in range(4, 10, 2):
        ws.row_dimensions[i].height = 24
        ws.row_dimensions[i + 1].height = 8
    ws["A11"] = "作業の流れ"; ws["A11"].font = F_LABEL
    flow = [
        "【毎年4月】 掲示用名簿を「名簿貼付」に貼る → ①解析 → ②反映（新入生は③登録）",
        "【支出のたび】 「支出入力」シートに件名と一人あたり金額 → ④一括入力 → 「支出承認書」を印刷",
        "【口座振替のたび】 「収入入力」シートに金額、未納者の精算番号だけ表に → ⑤一括入力",
        "【年度末】 ⑦決算集計 → 数字を決算書へ転記、⑧整合性チェック、⑨精算書PDF",
        "【年度の変わり目】 ⑬支出項目を読み込む → 業者名と○×を直す → ⑭業者別に集計 → ⑮○の項目だけ来年の予定表へ",
        "※どの書き込みも、実行前にマスターのバックアップが自動で作られます。",
    ]
    for i, t in enumerate(flow):
        ws.cell(row=12 + i, column=1, value=t).font = F_BODY
    for col, w in {"A": 3, "B": 26, "C": 3, "D": 3, "E": 26, "F": 3, "G": 3, "H": 26, "J": 3, "K": 26}.items():
        ws.column_dimensions[col].width = w

    # ============ 設定 ============
    ws = wb.create_sheet("設定")
    header(ws, "設定（最初に1回だけ入力）", 6)
    ic = label_input(ws, 3, "マスターファイルの場所",
                     note="例： C:\\Users\\jimu\\Desktop\\令和7年度生積立金.xlsx（フルパス）")
    ic = label_input(ws, 4, "バックアップの保存先",
                     note="空欄でOK（マスターと同じ場所に「バックアップ」フォルダを自動作成）")
    ic = label_input(ws, 5, "年度（数字）", note="例： 7　（令和7年度）")
    ic = label_input(ws, 6, "精算書PDFの保存先",
                     note="空欄でOK（マスターと同じ場所に「精算書PDF」フォルダを自動作成）")
    ic = label_input(ws, 7, "口座マスターの場所",
                     note="振替結果取込を使う場合のみ。例: C:\\Users\\jimu\\Desktop\\口座マスター.xlsx")
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 52
    ws["B8"] = "※黄色いセルが入力欄です。"
    ws["B8"].font = F_NOTE

    # ============ 名簿貼付 ============
    ws = wb.create_sheet("名簿貼付")
    ws["A1"] = "ここに掲示用名簿を貼り付け（シート全体をコピー → A1を選んで貼り付け）。貼ったら「メニュー」の①を実行。"
    ws["A1"].font = Font(name="游ゴシック", size=11, bold=True, color="C00000")

    # ============ 名簿一覧 ============
    ws = wb.create_sheet("名簿一覧")
    header(ws, "名簿一覧（①解析の結果）", 6)
    ws["A2"] = "D列が「見つからず」「複数候補」の行は、E列に正しい精算番号を手で入力してください。"
    ws["A2"].font = F_NOTE
    table_header(ws, 3, ["組", "番号", "氏名", "照合結果", "精算番号"])
    for col, w in {"A": 8, "B": 8, "C": 22, "D": 20, "E": 12}.items():
        ws.column_dimensions[col].width = w

    # ============ 支出入力 ============
    ws = wb.create_sheet("支出入力")
    header(ws, "支出の一括入力", 7)
    ws["A2"] = "黄色いセルを埋めて、メニューの「④支出をマスターへ一括入力」を実行します。"
    ws["A2"].font = F_NOTE
    ws["A3"] = "―基本情報―"; ws["A3"].font = F_LABEL
    label_input(ws, 4, "支出No（1〜100）", note="マスターのデータシート8行目のNo.。空いている番号を使う")
    label_input(ws, 5, "件名", note="例： 校外学習バス代")
    label_input(ws, 6, "支払先", note="例： ○○観光バス株式会社")
    label_input(ws, 7, "日付", note="例： 2026/5/20")
    label_input(ws, 8, "一人あたり金額（円）", note="例： 3500")
    label_input(ws, 9, "対象", note="「全員」か「例外表の生徒のみ」のどちらかを入力")
    ws["B11"] = "―例外表―　対象外の生徒は金額0、金額が違う生徒はその金額、返金はマイナスで"
    ws["B11"].font = F_LABEL
    ws["B12"] = "（例：転退学者→0、給付型奨学金の生徒→0、途中参加→半額など）"
    ws["B12"].font = F_NOTE
    table_header(ws, 13, ["精算番号", "組（メモ）", "番号（メモ）", "氏名（メモ）", "金額（0=対象外）"], start_col=2)
    for r in range(14, 1014):
        ws.cell(row=r, column=2).fill = FILL_IN
        ws.cell(row=r, column=6).fill = FILL_IN
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BORDER
    for col, w in {"A": 3, "B": 20, "C": 26, "D": 26, "E": 22, "F": 18}.items():
        ws.column_dimensions[col].width = w

    # ============ 収入入力 ============
    ws = wb.create_sheet("収入入力")
    header(ws, "収入（口座振替など）の一括入力", 7)
    ws["A2"] = "黄色いセルを埋めて、メニューの「⑤収入をマスターへ一括入力」を実行します。空き枠は「⑥収入枠の一覧を表示」で確認。"
    ws["A2"].font = F_NOTE
    label_input(ws, 4, "収入枠No（1〜43）", note="データシートJ列=枠1 〜 AZ列=枠43")
    label_input(ws, 5, "件名", note="例： マスター口座振替 10月分")
    label_input(ws, 6, "日付", note="例： 2026/10/13")
    label_input(ws, 7, "一人あたり金額（円）", note="例： 76000")
    ws["B9"] = "―未納者表―　振替できなかった生徒の精算番号だけを書く（他の生徒には自動で金額が入る）"
    ws["B9"].font = F_LABEL
    ws["B10"] = "ここに書いた生徒は空欄のままになり、マスターのH列に「未納」の印が自動で立ちます。"
    ws["B10"].font = F_NOTE
    table_header(ws, 11, ["精算番号", "組（メモ）", "番号（メモ）", "氏名（メモ）"], start_col=2)
    for r in range(12, 1012):
        ws.cell(row=r, column=2).fill = FILL_IN
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = BORDER
    for col, w in {"A": 3, "B": 20, "C": 26, "D": 26, "E": 22}.items():
        ws.column_dimensions[col].width = w

    # ============ 支出承認書 ============
    ws = wb.create_sheet("支出承認書")
    ws.merge_cells("B2:H2")
    c = ws["B2"]; c.value = "支　出　承　認　書"; c.font = Font(name="游ゴシック", size=16, bold=True)
    c.alignment = Alignment(horizontal="center")
    rows = [
        (4, "支出番号", "C4", "起案日", "H4"),
        (5, "年度（令和）", "C5", None, None),
        (6, "請求金額（円）", "C6", None, None),
        (7, "件名", "C7", None, None),
        (8, "支払先", "C8", None, None),
        (9, "対象生徒数（名）", "C9", None, None),
    ]
    for r, lbl, addr, lbl2, addr2 in rows:
        lc = ws.cell(row=r, column=2, value=lbl); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
        ws[addr].fill = FILL_OUT; ws[addr].border = BORDER
        if lbl2:
            lc2 = ws.cell(row=r, column=7, value=lbl2); lc2.font = F_LABEL; lc2.fill = FILL_LBL; lc2.border = BORDER
            ws[addr2].fill = FILL_OUT; ws[addr2].border = BORDER
    # 請求金額 = 品目表の合計
    ws["C6"] = "=F16"
    table_header(ws, 11, ["品目", "", "単価", "数量", "金額"], start_col=2)
    for r in range(12, 16):
        for cix in range(2, 7):
            ws.cell(row=r, column=cix).border = BORDER
        ws.cell(row=r, column=2).fill = FILL_OUT
        ws.cell(row=r, column=4).fill = FILL_OUT
        ws.cell(row=r, column=5).fill = FILL_OUT
        f = ws.cell(row=r, column=6)
        f.value = f'=IF(D{r}="","",D{r}*E{r})'
        f.fill = FILL_OUT
    lc = ws.cell(row=16, column=2, value="合計"); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
    ws["F16"] = "=SUM(F12:F15)"; ws["F16"].border = BORDER; ws["F16"].font = F_LABEL
    lc = ws.cell(row=17, column=2, value="一人あたり金額"); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
    ws["F17"] = '=IF(C9="","",ROUND(F16/C9,0))'; ws["F17"].border = BORDER
    lc = ws.cell(row=19, column=2, value="支払方法"); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
    ws["C19"] = "振込　・　現金"; ws["C19"].border = BORDER
    lc = ws.cell(row=20, column=2, value="支払期日"); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
    ws["C20"].fill = FILL_IN; ws["C20"].border = BORDER
    lc = ws.cell(row=21, column=2, value="備考"); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
    ws["C21"].fill = FILL_IN; ws["C21"].border = BORDER
    # 決裁欄
    table_header(ws, 23, ["校長", "副校長", "課(室)長", "事務担当", "学年等会計担当"], start_col=2)
    for cix in range(2, 7):
        ws.cell(row=24, column=cix).border = BORDER
    ws.row_dimensions[24].height = 42
    ws["B26"] = "※緑のセルは「④支出をマスターへ一括入力」実行時に自動で埋まります。黄色は手入力欄です。"
    ws["B26"].font = F_NOTE
    for col, w in {"A": 3, "B": 18, "C": 22, "D": 12, "E": 10, "F": 14, "G": 12, "H": 14}.items():
        ws.column_dimensions[col].width = w

    # ============ 収入承認書 ============
    ws = wb.create_sheet("収入承認書")
    ws.merge_cells("B2:H2")
    c = ws["B2"]; c.value = "収　入　承　認　書"; c.font = Font(name="游ゴシック", size=16, bold=True)
    c.alignment = Alignment(horizontal="center")
    fields = [
        (4, "収入番号", FILL_IN),
        (5, "年度（令和）", FILL_OUT),
        (6, "収入金額（円）", FILL_OUT),
        (7, "収入区分", FILL_IN),
        (8, "収入年月日", FILL_OUT),
        (9, "件名", FILL_OUT),
        (10, "納入者", FILL_IN),
        (11, "内訳", FILL_IN),
        (12, "備考", FILL_IN),
    ]
    for r, lbl, fill in fields:
        lc = ws.cell(row=r, column=2, value=lbl); lc.font = F_LABEL; lc.fill = FILL_LBL; lc.border = BORDER
        ic = ws.cell(row=r, column=3); ic.fill = fill; ic.border = BORDER
    ws["C7"] = "マスター口座からの振替"
    table_header(ws, 14, ["校長", "副校長", "課(室)長", "事務担当"], start_col=2)
    for cix in range(2, 6):
        ws.cell(row=15, column=cix).border = BORDER
    ws.row_dimensions[15].height = 42
    ws["B17"] = "※緑のセルは「⑤収入をマスターへ一括入力」実行時に自動で埋まります。黄色は手入力欄です。"
    ws["B17"].font = F_NOTE
    for col, w in {"A": 3, "B": 18, "C": 32, "D": 12, "E": 12}.items():
        ws.column_dimensions[col].width = w

    # ============ 決算集計 ============
    ws = wb.create_sheet("決算集計")
    header(ws, "決算集計（⑦の結果がここに出ます）", 7)
    ws["A2"] = "この数字を決算書（1年決算書シート）へ転記してください。「一人あたり」は最も多くの生徒に入っている金額です。"
    ws["A2"].font = F_NOTE
    table_header(ws, 3, ["区分", "No", "件名", "日付", "対象人数", "一人あたり", "執行総額"])
    for col, w in {"A": 8, "B": 8, "C": 34, "D": 12, "E": 10, "F": 12, "G": 14}.items():
        ws.column_dimensions[col].width = w

    # ============ 振替結果取込 ============
    ws = wb.create_sheet("振替結果取込")
    header(ws, "振替結果取込（銀行の結果を貼るだけで未納者表を自動作成）", 9)
    ws["A2"] = "銀行の振替結果の 口座記号・口座番号・金額・振替結果 を12行目から貼り付け、メニューの「⑪振替結果を照合」を実行します。"
    ws["A2"].font = F_NOTE
    ws["A3"] = "照合には「設定」シートC7の口座マスターを使います。結果はG〜I列と、収入入力シートの未納者表に自動で入ります。"
    ws["A3"].font = F_NOTE
    # サマリー欄（自動出力）
    for r, label in [(5, "読取件数"), (6, "うち振替済"), (7, "うち未納"), (8, "不明口座（要確認）"), (9, "要確認（口座重複・重複行）")]:
        lc = ws.cell(row=r, column=7, value=label)
        lc.font = F_LABEL
        lc.fill = FILL_LBL
        lc.border = BORDER
        oc = ws.cell(row=r, column=8)
        oc.fill = FILL_OUT
        oc.border = BORDER
    table_header(ws, 11, ["口座記号", "口座番号", "金額", "振替結果"], start_col=2)
    table_header(ws, 11, ["精算番号(自動)", "氏名(自動)", "判定(自動)"], start_col=7)
    for r in range(12, 1012):
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_IN
            cell.border = BORDER
        for c in range(7, 10):
            ws.cell(row=r, column=c).border = BORDER
    for col, w in {"A": 3, "B": 12, "C": 12, "D": 12, "E": 14, "F": 3, "G": 14, "H": 18, "I": 16}.items():
        ws.column_dimensions[col].width = w

    # ============ 年間予定表 ============
    ws = wb.create_sheet("年間予定表")
    header(ws, "年間徴収・支出予定表（年度初めに計画を書き、実行時は行番号で呼び出す）", 8)
    ws["A2"] = "実行するときはメニューの「⑫予定を入力フォームへ転送」で行番号を指定 → 支出入力/収入入力シートに自動転記されます。"
    ws["A2"].font = F_NOTE
    table_header(ws, 3, ["行No", "予定月", "区分(支出/収入)", "No(支出No/収入枠No)", "件名", "支払先", "一人あたり金額", "メモ"])
    for i, r in enumerate(range(4, 64)):
        nc = ws.cell(row=r, column=1, value=i + 1)
        nc.border = BORDER
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_IN
            cell.border = BORDER
    for col, w in {"A": 6, "B": 8, "C": 14, "D": 16, "E": 28, "F": 20, "G": 14, "H": 20}.items():
        ws.column_dimensions[col].width = w

    # ============ 支出項目一覧 ============
    ws = wb.create_sheet("支出項目一覧")
    header(ws, "支出項目一覧（業者ごとの見える化と、来年度への引き継ぎ）", 9)
    ws["A2"] = "メニューの「⑬支出項目を読み込む」でマスターの支出100枠がここに並びます（マスターは読むだけ・書き換えません）。"
    ws["A2"].font = F_NOTE
    ws["A3"] = "C列の業者名は自由に編集OK（同じ業者名＝⑭でひとつに合算）。来年も使う項目はH列に○、使わないなら×。○だけが⑮で年間予定表に引き継がれます。"
    ws["A3"].font = F_NOTE
    table_header(ws, 5, ["支出No", "件名（マスターから）", "業者名（←自由に編集）", "日付",
                         "対象人数", "一人あたり", "執行総額", "来年も使う(○/×)", "メモ"])
    for r in range(6, 106):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c in (3, 8, 9):          # 業者名・○×・メモは手で直せる欄
                cell.fill = FILL_IN
            else:                        # それ以外は⑬が自動で埋める欄
                cell.fill = FILL_OUT
    # 右側：⑭業者別に集計する の出力欄（K〜Y列）
    ws["K4"] = "―業者別集計（⑭の結果がここに出ます）―"
    ws["K4"].font = F_LABEL
    months = [f"{m}月" for m in list(range(4, 13)) + list(range(1, 4))]
    table_header(ws, 5, ["業者名", "項目数", "年間合計"] + months, start_col=11)
    for r in range(6, 66):
        for c in range(11, 26):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.fill = FILL_OUT
    widths = {"A": 8, "B": 30, "C": 24, "D": 12, "E": 9, "F": 11, "G": 13, "H": 15, "I": 16, "J": 3,
              "K": 24, "L": 8, "M": 13}
    for i in range(12):
        widths[chr(ord("N") + i)] = 10
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ============ チェック結果 ============
    ws = wb.create_sheet("チェック結果")
    header(ws, "整合性チェックの結果（⑧の結果がここに出ます）", 6)
    for col, w in {"A": 60, "B": 14, "C": 18, "D": 12}.items():
        ws.column_dimensions[col].width = w

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, "積立金入力アシスタント.xlsx")
    wb.save(out)
    print("saved", out)


if __name__ == "__main__":
    build()
