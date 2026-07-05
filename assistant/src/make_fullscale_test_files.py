# -*- coding: utf-8 -*-
"""フルスケール検証用の擬似データ一式を生成するスクリプト

お預かりした実ファイルの構造（8クラス×40名・掲示用名簿のブロック座標・
マスターの行列配置）をそのまま再現した、架空の氏名320名のデータセット。
本番と同じ規模・同じ配置で入力動作を検証するためのもの。

実物の掲示用名簿の幾何配置（解析結果に一致させてある）:
  ・クラス見出し: 2行目、D2から10列間隔（D2,N2,X2,AH2,AR2,BB2,BL2,BV2）
  ・番号列 = 見出しの1列左（C,M,W,AG,AQ,BA,BK,BU）
  ・氏名列 = 見出しの2列右（F,P,Z,AJ,AT,BD,BN,BX）
  ・生徒行 = 4行目から40行

生成物（assistant/output/）:
  検証用_掲示用名簿.xlsx        … 1年生のクラス発表名簿（8クラス×40名）
  検証用_新年度名簿.xlsx        … 同じ320名をシャッフルした2年生名簿（クラス替え検証用）
  検証用_令和X年度生積立金.xlsx … 320名が登録済みのフルスケールマスター
  検証用_口座マスター.xlsx      … 320名分の架空口座
  検証用_振替結果.xlsx          … 320件+架空1件（未納5名・不明口座1件入り）
"""
import os
import random

from make_practice_files import SEI, MEI, build_master

from openpyxl import Workbook
from openpyxl.styles import Font

random.seed(320)

N_CLASS = 8
N_PER_CLASS = 40
N = N_CLASS * N_PER_CLASS  # 320

# 検証シナリオ: 未納にする生徒（0始まりの通し番号）と不明口座の有無
UNPAID = (6, 43, 158, 240, 311)   # 精算番号 7, 44, 159, 241, 312


def fake_names(n):
    """架空のひらがな氏名をn人分（重複なし・五十音順ではなくランダム順で現実感を出す）"""
    names = set()
    while len(names) < n:
        names.add(f"{random.choice(SEI)}　{random.choice(MEI)}")
    names = sorted(names)
    random.shuffle(names)
    return names


def build_roster_realgeom(assignments, out, grade, title_note):
    """実物と同一の幾何配置で掲示用名簿を作る。
    assignments: [(組, 番号, 氏名), ...]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "掲示用"
    ws["B1"] = title_note
    ws["B1"].font = Font(size=9, color="808080")
    by_class = {}
    for kumi, ban, name in assignments:
        by_class.setdefault(kumi, []).append((ban, name))
    for k in range(1, N_CLASS + 1):
        head_col = 4 + (k - 1) * 10           # D=4, N=14, ... BV=74
        hdr = ws.cell(row=2, column=head_col, value=f"{grade}年{k}組")
        hdr.font = Font(bold=True, size=14)
        ws.cell(row=3, column=head_col - 1, value="担任")
        for ban, name in sorted(by_class.get(k, [])):
            r = 3 + ban                        # 番号1 → 4行目
            ws.cell(row=r, column=head_col - 1, value=ban)      # 番号列（見出し-1）
            ws.cell(row=r, column=head_col + 2, value=name)     # 氏名列（見出し+2）
    wb.save(out)
    print("saved", out)


def build_account_master_full(names, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "口座マスター"
    ws["A1"] = "口座マスター（検証用・架空の口座番号320名）"
    ws["A1"].font = Font(bold=True, size=14)
    for c, h in enumerate(["精算番号", "氏名", "口座記号(5桁)", "口座番号(8桁)", "備考"], start=1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)
    for i, name in enumerate(names):
        r = 4 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=f"{10000 + i:05d}")
        ws.cell(row=r, column=4, value=f"{80000000 + i * 111:08d}")
    for col, w in {"A": 10, "B": 20, "C": 14, "D": 14, "E": 16}.items():
        ws.column_dimensions[col].width = w
    wb.save(out)
    print("saved", out)


def build_bank_result_full(names, out):
    """320件の振替結果 + 口座マスターに存在しない1件（不明口座の動作確認用）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "振替結果"
    ws["A1"] = "自動払込み 振替結果票（検証用・架空320件+口座不一致1件）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "末尾の1件はわざと口座マスターに無い口座にしてある（⑪で「不明口座」と出るのが正常動作）。"
    for c, h in enumerate(["口座記号", "口座番号", "金額", "振替結果"], start=1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)
    r = 4
    for i in range(len(names)):
        ws.cell(row=r, column=1, value=f"{10000 + i:05d}")
        ws.cell(row=r, column=2, value=f"{80000000 + i * 111:08d}")
        if i in UNPAID:
            ws.cell(row=r, column=3, value=0)
            ws.cell(row=r, column=4, value="1")     # 資金不足
        else:
            ws.cell(row=r, column=3, value=76000)
            ws.cell(row=r, column=4, value="0")     # 振替済
        r += 1
    # 不明口座（口座マスターに存在しない）
    ws.cell(row=r, column=1, value="99999")
    ws.cell(row=r, column=2, value="12345678")
    ws.cell(row=r, column=3, value=76000)
    ws.cell(row=r, column=4, value="0")
    for col, w in {"A": 12, "B": 14, "C": 12, "D": 12}.items():
        ws.column_dimensions[col].width = w
    wb.save(out)
    print("saved", out)


def main():
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
    os.makedirs(out_dir, exist_ok=True)

    names = fake_names(N)

    # 1年時の割当: 通し番号順に 1組1番, 1組2番, ...（精算番号=登録順と一致させる）
    first_year = [(i // N_PER_CLASS + 1, i % N_PER_CLASS + 1, name)
                  for i, name in enumerate(names)]
    build_roster_realgeom(first_year, os.path.join(out_dir, "検証用_掲示用名簿.xlsx"),
                          grade=1, title_note="検証用ダミー（架空の氏名320名・実物と同じ配置）")

    # 2年時の割当: 全員をシャッフルして新しい組・番号に（クラス替え①②の検証用）
    shuffled = names[:]
    random.shuffle(shuffled)
    second_year = [(i // N_PER_CLASS + 1, i % N_PER_CLASS + 1, name)
                   for i, name in enumerate(shuffled)]
    build_roster_realgeom(second_year, os.path.join(out_dir, "検証用_新年度名簿.xlsx"),
                          grade=2, title_note="検証用ダミー（同じ320名のクラス替え後・①②の検証用）")

    # フルスケールマスター（make_practice_files.build_master を320名構成で流用）
    import make_practice_files as mp
    mp.N_PER_CLASS = N_PER_CLASS   # 組・番号の割当を8クラス×40名に合わせる
    build_master(names, os.path.join(out_dir, "検証用_令和X年度生積立金.xlsx"), filled=True)

    build_account_master_full(names, os.path.join(out_dir, "検証用_口座マスター.xlsx"))
    build_bank_result_full(names, os.path.join(out_dir, "検証用_振替結果.xlsx"))

    print()
    print(f"検証データ一式（{N}名構成）の生成が完了。")
    print(f"想定結果: ⑪で 読取{N + 1}件／振替済{N - len(UNPAID)}名／未納{len(UNPAID)}名／不明口座1件")
    print(f"未納の精算番号: {[i + 1 for i in UNPAID]}")


if __name__ == "__main__":
    main()
