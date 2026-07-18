# -*- coding: utf-8 -*-
"""製品版パッケージ（学校納品用の完成形）を生成する。

日本の業務ソフトのパッケージ構成に倣い、本体・かんたん導入・練習データ・
マニュアル・動作確認・ライセンス・リリースノートを1つにまとめる。
デモキット（営業用）・導入用USB（郵送用）とは別の「納品物そのもの」。

生成先: assistant/output/製品パッケージ/ と 積立金入力アシスタント_v{VERSION}.zip

実行前提: build_assistant.py / make_practice_files.py /
          make_fullscale_test_files.py / docs各PDF が生成済みであること。

★出荷前に必ずやること（Windows作業・1回だけ）
  01_本体/積立金入力アシスタント.xlsx に A00〜A11.bas をインポートして
  .xlsm 保存したものに差し替え、「★出荷前に差し替え.txt」と
  「VBAモジュール予備」を削除してから納品する。
"""
import os
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from build_assistant import VERSION

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
OUT = os.path.join(REPO, "assistant", "output")
PKG = os.path.join(OUT, "製品パッケージ")

NAVY = "203864"
YELLOW = "FFF3B0"
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_text(path, text):
    with open(path, "w", encoding="utf-8-sig", newline="\r\n") as f:
        f.write(text)


def copy(rel, dst, newname=None, required=True):
    src = os.path.join(REPO, rel)
    if not os.path.exists(src):
        if required:
            raise FileNotFoundError(src)
        print(f"  - skip (not found): {rel}")
        return
    name = newname or os.path.basename(src)
    if os.path.isdir(src):
        shutil.copytree(src, os.path.join(dst, name))
    else:
        shutil.copy2(src, os.path.join(dst, name))
    print(f"  + {os.path.relpath(os.path.join(dst, name), OUT)}")


# ------------------------------------------------------------------
# 同梱物1: 口座マスターひな形（空のテンプレート）
# ------------------------------------------------------------------
def make_account_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "口座マスター"
    ws["A1"] = "口座マスター（精算番号と、ゆうちょ口座の対応表）"
    ws["A1"].font = Font(name="游ゴシック", size=14, bold=True, color=NAVY)
    ws["A2"] = "4行目から、生徒の精算番号・氏名・口座記号・口座番号を入力してください。⑪振替結果の照合で使います。"
    ws["A2"].font = Font(name="游ゴシック", size=10, color="808080")
    headers = ["精算番号", "氏名", "口座記号", "口座番号"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
    for r in range(4, 404):
        for j in range(1, 5):
            cell = ws.cell(row=r, column=j)
            cell.fill = PatternFill("solid", fgColor=YELLOW)
            cell.border = BORDER
    for col, w in {"A": 10, "B": 22, "C": 12, "D": 14}.items():
        ws.column_dimensions[col].width = w
    wb.save(path)
    print(f"  + {os.path.relpath(path, OUT)}")


# ------------------------------------------------------------------
# 同梱物2: 導入前チェック（マクロを一切使わない環境確認ブック）
# ------------------------------------------------------------------
def make_precheck(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "導入前チェック"
    ws["A1"] = "導入前チェック（このファイルにマクロは入っていません。開くだけで安全です）"
    ws["A1"].font = Font(name="游ゴシック", size=13, bold=True, color=NAVY)

    ws["A3"] = "1. お使いのExcelのバージョン（自動表示）"
    ws["A3"].font = Font(name="游ゴシック", size=11, bold=True)
    ws["B4"] = '=IF(ISERROR(INFO("release")),"表示できませんでした（古いExcelの可能性があります）","Excel バージョン " & INFO("release"))'
    ws["B5"] = "→ 「16.0」以上と表示されれば動作環境（Excel 2016以降）を満たしています。"
    ws["B5"].font = Font(name="游ゴシック", size=10, color="808080")

    ws["A7"] = "2. 手で確認する項目（済んだらB列に○）"
    ws["A7"].font = Font(name="游ゴシック", size=11, bold=True)
    checks = [
        "OSは Windows 10 または 11 である",
        "このパソコンに Microsoft Excel が入っている（互換ソフトではない）",
        "デスクトップに新しいフォルダを作成できる（書き込みが禁止されていない）",
        "「かんたん導入ガイド.pdf」を印刷できる（または画面で見られる）",
    ]
    for i, t in enumerate(checks):
        r = 8 + i
        c = ws.cell(row=r, column=2)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = BORDER
        ws.cell(row=r, column=3, value=t).font = Font(name="游ゴシック", size=11)

    ws["A13"] = "3. すべて確認できたら"
    ws["A13"].font = Font(name="游ゴシック", size=11, bold=True)
    ws["B14"] = "「かんたん導入ガイド.pdf」の手順1へ進んでください。"
    ws["B14"].font = Font(name="游ゴシック", size=11)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 70
    wb.save(path)
    print(f"  + {os.path.relpath(path, OUT)}")


# ------------------------------------------------------------------
# 同梱テキスト
# ------------------------------------------------------------------
README_ROOT = f"""＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝
　積立金会計 入力アシスタント　バージョン {VERSION}
＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝

このパッケージは、学校徴収金（積立金）の入力業務を自動化する
Excelブック製品です。専用ソフトのインストールはありません。

【導入は3ステップです】
　1.「02_かんたん導入」の 導入前チェック.xlsx を開いて環境を確認
　2.「01_本体」フォルダを丸ごとデスクトップにコピー
　3. かんたん導入ガイド.pdf の手順どおりにマクロを許可して、開く

【同梱物】
　00_はじめにお読みください.txt … この文書
　01_本体/                      … 製品本体（Excelブック）と口座マスターひな形
　02_かんたん導入/              … 導入前チェック・導入ガイド・管理者様向けご説明
　03_練習用データ/              … 架空の80名の練習環境（実在の生徒情報は不使用）
　04_マニュアル/                … 使い方ガイド（図解）・機能別手順書
　05_動作確認/                  … 動作確認チェックシートと検証用データ（架空の320名）
　ライセンス.txt                … 利用条件
　リリースノート.txt            … 変更履歴

【はじめての方へ】
　いきなり本番のファイルには使いません。まず 03_練習用データ で
　ひととおり練習し、05_動作確認 のチェックシートが全部○になってから
　本番のマスターファイルに切り替えてください。
"""

LICENSE = f"""積立金会計 入力アシスタント バージョン {VERSION}　利用条件

1. 本製品は、ご契約いただいた学校（1校）の校務のためにご利用いただけます。
2. 契約校の校内でのコピー（バックアップ・複数PCへの配置）は自由です。
   契約校以外への配布・譲渡・転売はできません。
3. 本製品の改変は、サポートの対象外となる場合があります。
   様式変更などのご要望は改修として承ります。
4. 本製品は現状有姿で提供されます。導入時は必ず同梱の練習用データで
   動作確認を行ってから実データにお使いください。
5. 本製品の利用により生じた損害について、提供者は故意または重過失に
   よる場合を除き、責任を負いません（書き込み前の自動バックアップに
   より、実行前の状態にいつでも戻せます）。
6. ご契約の解除後も、作成済みのデータ・マスターファイルはそのまま
   お使いいただけます（本製品によるデータの囲い込みはありません）。
"""

RELEASE_NOTES = f"""リリースノート

■ バージョン {VERSION}（初回リリース）
・名簿の自動解析・照合、クラス替えの一括反映、新入生登録（①〜③）
・支出・収入の一括入力と承認書の自動作成（④〜⑥）
・決算集計、整合性チェック、精算書の一括PDF/一括印刷（⑦〜⑩）
・銀行振替結果の自動照合（口座マスター連携・⑪）
・年間予定表からの入力フォーム自動転記（⑫）
・支出項目の一覧化・業者別集計・翌年度への引き継ぎ（⑬〜⑮）
・安全装置：書き込み前の自動バックアップ／マスター構造チェック／
　上書き前の確認／数式列・名簿列への書き込み禁止
・検証：実物と同形式のファイルおよび架空320名データで全機能を確認。
　異常データ15項目の耐性テストを実施し、判明した問題はすべて修正済み
"""


def main():
    if os.path.exists(PKG):
        shutil.rmtree(PKG)
    os.makedirs(PKG)

    write_text(os.path.join(PKG, "00_はじめにお読みください.txt"), README_ROOT)
    write_text(os.path.join(PKG, "ライセンス.txt"), LICENSE)
    write_text(os.path.join(PKG, "リリースノート.txt"), RELEASE_NOTES)

    # ---- 01 本体 ----
    d1 = os.path.join(PKG, "01_本体")
    os.makedirs(d1)
    copy("assistant/output/積立金入力アシスタント.xlsx", d1)
    make_account_template(os.path.join(d1, "口座マスターひな形.xlsx"))
    write_text(os.path.join(d1, "★出荷前に差し替え.txt"),
               "【提供者側の作業メモ・納品前に削除】\r\n"
               "この .xlsx に VBAモジュール予備/A00〜A11.bas をインポートし\r\n"
               "「積立金入力アシスタント.xlsm」として保存したものに差し替える。\r\n"
               "差し替え後、このメモと「VBAモジュール予備」フォルダを削除して納品する。\r\n")
    vba_dst = os.path.join(d1, "VBAモジュール予備")
    os.makedirs(vba_dst)
    vba_src = os.path.join(REPO, "assistant", "vba")
    for name in sorted(os.listdir(vba_src)):
        if name.endswith(".bas"):
            shutil.copy2(os.path.join(vba_src, name), vba_dst)
    print("  + 01_本体/VBAモジュール予備/(12本)")

    # ---- 02 かんたん導入 ----
    d2 = os.path.join(PKG, "02_かんたん導入")
    os.makedirs(d2)
    make_precheck(os.path.join(d2, "導入前チェック.xlsx"))
    copy("docs/04_オンライン導入/オンライン導入手順書.pdf", d2, "かんたん導入ガイド.pdf")
    copy("docs/01_学校向けマニュアル/学校管理者様向けご説明.pdf", d2)

    # ---- 03 練習用データ ----
    d3 = os.path.join(PKG, "03_練習用データ")
    os.makedirs(d3)
    for f in ["練習用_令和X年度生積立金.xlsx", "練習用_口座マスター.xlsx",
              "練習用_振替結果.xlsx", "練習用_掲示用名簿.xlsx", "練習用_空のマスター.xlsx"]:
        copy(f"assistant/output/{f}", d3)
    write_text(os.path.join(d3, "練習データについて.txt"),
               "架空の80名（4クラス×20名）の練習環境です。実在の生徒・口座情報は\r\n"
               "一切含まれません。設定シートのC3・C7をこのフォルダのファイルに\r\n"
               "向ければ、何度でも安全に練習できます。\r\n")

    # ---- 04 マニュアル ----
    d4 = os.path.join(PKG, "04_マニュアル")
    os.makedirs(d4)
    copy("docs/01_学校向けマニュアル/使い方ガイド_図解版.pdf", d4)
    copy("docs/01_学校向けマニュアル/入力アシスタント_手順書.pdf", d4)

    # ---- 05 動作確認 ----
    d5 = os.path.join(PKG, "05_動作確認")
    os.makedirs(d5)
    copy("docs/01_学校向けマニュアル/動作確認チェックシート.pdf", d5)
    dv = os.path.join(d5, "検証用データ_320名")
    os.makedirs(dv)
    for f in ["検証用_令和X年度生積立金.xlsx", "検証用_口座マスター.xlsx",
              "検証用_振替結果.xlsx", "検証用_掲示用名簿.xlsx", "検証用_新年度名簿.xlsx"]:
        copy(f"assistant/output/{f}", dv)

    # ---- ZIP化 ----
    zip_path = os.path.join(OUT, f"積立金入力アシスタント_v{VERSION}.zip")
    if os.path.exists(zip_path):
        os.remove(zip_path)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(PKG):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.relpath(fp, OUT))
    print(f"\n完成: {PKG}")
    print(f"完成: {zip_path} ({os.path.getsize(zip_path)/1024/1024:.1f} MB)")


if __name__ == "__main__":
    main()
