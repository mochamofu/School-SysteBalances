# -*- coding: utf-8 -*-
"""実機テストキット（2台PC・データ反映検証用）を生成する。

目的: 手持ちのノートPC2台で「USBで渡したデータが、マスターに
正しく反映されるか」を自分の手で確かめるための一式。
学校導入の前段階（自分による最終検証）に使う。

デモキット（営業用）・製品パッケージ（納品用）とは別物で、
こちらは「テストして記録する」ことに特化している。

生成物:
  assistant/output/実機テストキット/
    00_テストのはじめかた.txt
    01_PC1_記録側/           … アシスタント＋VBA＋テストデータ
    02_PC2_送信側/           … 送信側が受け取る想定のデータ
    03_記録用紙/             … テスト結果記録ブック.xlsx＋手順書PDF
  assistant/output/実機テストキット.zip
"""
import os
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
OUT = os.path.join(REPO, "assistant", "output")
KIT = os.path.join(OUT, "実機テストキット")

NAVY = "203864"
YELLOW = "FFF3B0"   # 入力欄
GREEN = "E2F6E6"    # 自動判定
GRAY = "F2F2F2"

F_TITLE = Font(name="游ゴシック", size=15, bold=True, color=NAVY)
F_H = Font(name="游ゴシック", size=10.5, bold=True, color="FFFFFF")
F_LBL = Font(name="游ゴシック", size=10.5, bold=True, color=NAVY)
F_BODY = Font(name="游ゴシック", size=10.5)
F_NOTE = Font(name="游ゴシック", size=9.5, color="808080")

FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_IN = PatternFill("solid", fgColor=YELLOW)
FILL_OUT = PatternFill("solid", fgColor=GREEN)
FILL_GRAY = PatternFill("solid", fgColor=GRAY)

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def write_text(path, text):
    with open(path, "w", encoding="utf-8-sig", newline="\r\n") as f:
        f.write(text)


def copy(rel, dst, newname=None, required=True):
    src = os.path.join(REPO, rel)
    if not os.path.exists(src):
        if required:
            raise FileNotFoundError(src)
        print(f"  - skip: {rel}")
        return
    name = newname or os.path.basename(src)
    if os.path.isdir(src):
        shutil.copytree(src, os.path.join(dst, name))
    else:
        shutil.copy2(src, os.path.join(dst, name))
    print(f"  + {os.path.relpath(os.path.join(dst, name), OUT)}")


def head_row(ws, row, cols, widths=None):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = F_H
        cell.fill = FILL_H
        cell.border = BORDER
        cell.alignment = CENTER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ==================================================================
# テスト結果記録ブック（マクロなし・数式で自動判定）
# ==================================================================
def build_record_book(path):
    wb = Workbook()

    # ---------- 1. 実施情報 ----------
    ws = wb.active
    ws.title = "① 実施情報"
    ws["A1"] = "実機テスト 実施情報"
    ws["A1"].font = F_TITLE
    ws["A2"] = "黄色いセルを埋めてください。テストの記録として保管します（融資面談や学校への提示にも使えます）。"
    ws["A2"].font = F_NOTE

    items = [
        ("実施日", "例: 2026/8/10"),
        ("実施者", ""),
        ("PC①（記録側）の機種名", "例: 富士通 LIFEBOOK"),
        ("PC①のWindowsバージョン", "設定→システム→バージョン情報"),
        ("PC①のExcelバージョン", "Excel→ファイル→アカウント→Excelのバージョン情報"),
        ("PC②（送信側）の機種名", ""),
        ("PC②のWindowsバージョン", ""),
        ("PC②のExcelバージョン", "PC②はExcelなしでも可（データを運ぶだけ）"),
        ("開始時刻", ""),
        ("終了時刻", ""),
    ]
    for i, (label, note) in enumerate(items):
        r = 4 + i
        c = ws.cell(row=r, column=1, value=label)
        c.font = F_LBL; c.fill = PatternFill("solid", fgColor="DCE8FA"); c.border = BORDER
        ic = ws.cell(row=r, column=2)
        ic.fill = FILL_IN; ic.border = BORDER; ic.font = F_BODY
        nc = ws.cell(row=r, column=3, value=note)
        nc.font = F_NOTE
    for col, w in {"A": 28, "B": 26, "C": 46}.items():
        ws.column_dimensions[col].width = w

    ws["A16"] = "総合判定（②〜③のシートを埋めると自動で出ます）"
    ws["A16"].font = F_LBL
    ws["B16"] = ('=IF(COUNTIF(\'② 動作テスト\'!F:F,"×")+COUNTIF(\'③ データ反映の突合\'!E:E,"×")>0,'
                 '"要修正（×あり）",IF(COUNTA(\'② 動作テスト\'!E5:E24)=0,"未実施","全項目OK"))')
    ws["B16"].fill = FILL_OUT; ws["B16"].border = BORDER
    ws["B16"].font = Font(name="游ゴシック", size=12, bold=True, color="1E7A3C")

    # ---------- 2. 動作テスト ----------
    ws = wb.create_sheet("② 動作テスト")
    ws["A1"] = "動作テスト ― 上から順に実行し、実測値（E列）を入力してください"
    ws["A1"].font = F_TITLE
    ws["A2"] = "判定（F列）は自動で出ます。×が出たら「④ 不具合記録」に状況を書いてください。"
    ws["A2"].font = F_NOTE

    head_row(ws, 4,
             ["No", "テスト内容", "操作", "期待値", "実測値（入力）", "判定"],
             [5, 26, 40, 22, 18, 8])

    # (テスト名, 操作, 期待値, 期待値の型)  数値比較のものは数値で
    tests = [
        ("導入", "VBA12本をインポート→Alt+F8→初期設定", "ボタンの数", 15),
        ("導入", "積立金入力アシスタント.xlsm を開く", "警告バーなしで開けたら1", 1),
        ("練習①", "名簿貼付に練習用_掲示用名簿を貼る→①名簿を解析", "検出したクラス数", 4),
        ("練習①", "同上", "検出した人数", 80),
        ("練習①", "同上", "「一致」以外の行数", 0),
        ("練習②", "②クラス替えをマスターに反映", "更新された人数", 80),
        ("練習③", "支出入力に 件名=校外学習バス代 金額=3500 対象=全員→④", "対象生徒数", 80),
        ("練習③", "同上", "合計金額（円）", 280000),
        ("練習④", "振替結果取込B12に練習用_振替結果の4列×80行を貼る→⑪", "読取件数", 80),
        ("練習④", "同上", "振替済", 78),
        ("練習④", "同上", "未納", 2),
        ("練習④", "同上", "不明口座", 0),
        ("練習⑤", "収入入力シートを見る（⑪の直後）", "未納者表に入った人数", 2),
        ("練習⑤", "収入入力に 金額=76000→⑤収入をマスターへ一括入力", "入金あり人数", 78),
        ("練習⑥", "⑥収入枠の一覧を表示", "使った枠の人数表示", 78),
        ("練習⑦", "⑦決算用の集計を実行", "決算集計シートに並んだ項目数", 2),
        ("練習⑧", "⑧マスターの整合性をチェック", "重大な不整合の件数", 0),
        ("検証①", "設定C3/C7を検証用に変更→検証用_新年度名簿→①", "検出した人数", 320),
        ("検証②", "検証用_振替結果（321行）を貼る→⑪", "読取件数", 321),
        ("検証②", "同上", "未納", 5),
    ]
    for i, (name, op, exp_label, exp_val) in enumerate(tests):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1).border = BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        for col, val in ((2, name), (3, op), (4, exp_label)):
            c = ws.cell(row=r, column=col, value=val)
            c.font = F_BODY; c.border = BORDER; c.alignment = WRAP
        # 期待値を D列の隣（H列）に隠し持たせず、期待値セル自体を数値で置く
        ec = ws.cell(row=r, column=4)
        ec.value = f"{exp_label} = {exp_val}"
        # 実測値入力
        ic = ws.cell(row=r, column=5)
        ic.fill = FILL_IN; ic.border = BORDER; ic.font = F_BODY; ic.alignment = CENTER
        # 判定（H列に期待値の数値を置いて比較）
        hv = ws.cell(row=r, column=8, value=exp_val)
        hv.font = F_NOTE
        jc = ws.cell(row=r, column=6,
                     value=f'=IF(E{r}="","",IF(E{r}=H{r},"○","×"))')
        jc.fill = FILL_OUT; jc.border = BORDER; jc.alignment = CENTER
        jc.font = Font(name="游ゴシック", size=12, bold=True)
    ws.column_dimensions["H"].width = 10
    ws.cell(row=4, column=8, value="（期待値・自動判定用）").font = F_NOTE
    ws.freeze_panes = "A5"

    # ---------- 3. データ反映の突合 ----------
    ws = wb.create_sheet("③ データ反映の突合")
    ws["A1"] = "データ反映の突合 ― マスターの中身が本当に変わったかをセル単位で確認"
    ws["A1"].font = F_TITLE
    ws["A2"] = ("練習用マスター（練習用_令和X年度生積立金.xlsx）の「データ」シートを開き、"
                "下の「見る場所」のセルの値をD列に書き写してください。")
    ws["A2"].font = F_NOTE
    ws["A3"] = "※行の見方: 精算番号1 = 9行目、精算番号7 = 15行目、精算番号44 = 52行目、合計行 = 331行目"
    ws["A3"].font = F_NOTE

    head_row(ws, 5,
             ["確認したいこと", "見る場所（セル）", "期待される値", "実際の値（入力）", "判定"],
             [34, 20, 22, 18, 8])

    checks = [
        ("④支出：1人目に金額が入ったか", "BE9", 3500),
        ("④支出：2人目にも入ったか", "BE10", 3500),
        ("④支出：最後の生徒(80人目)にも入ったか", "BE88", 3500),
        ("④支出：合計行が正しいか", "BE331", 280000),
        ("④支出：支出合計の列が反応したか", "FC9", 3500),
        ("⑤収入：1人目に入金額が入ったか", "J9", 76000),
        ("⑤収入：未納者(精算番号7)は空欄のままか", "J15", 0),
        ("⑤収入：未納者(精算番号44)は空欄のままか", "J52", 0),
        ("⑤収入：収入合計の列が反応したか", "BC9", 76000),
        ("②クラス替え：1人目の組が更新されたか", "E9", 1),
        ("②クラス替え：精算番号の並びが変わっていないか", "A9", 1),
        ("②クラス替え：氏名が書き換わっていないか", "G9", 0),
    ]
    for i, (what, cellref, exp) in enumerate(checks):
        r = 6 + i
        c1 = ws.cell(row=r, column=1, value=what); c1.font = F_BODY; c1.border = BORDER; c1.alignment = WRAP
        c2 = ws.cell(row=r, column=2, value=cellref); c2.font = F_LBL; c2.border = BORDER; c2.alignment = CENTER
        if what.endswith("空欄のままか"):
            lbl = "空欄（何も入っていない）"
        elif "氏名が書き換わっていない" in what:
            lbl = "テスト前と同じ氏名"
        elif "並びが変わっていない" in what:
            lbl = "1（変わらない）"
        elif "組が更新された" in what:
            lbl = "新しい組の番号"
        else:
            lbl = f"{exp:,}"
        c3 = ws.cell(row=r, column=3, value=lbl); c3.font = F_BODY; c3.border = BORDER; c3.alignment = WRAP
        ic = ws.cell(row=r, column=4); ic.fill = FILL_IN; ic.border = BORDER; ic.alignment = CENTER
        hv = ws.cell(row=r, column=7, value=exp); hv.font = F_NOTE
        # 空欄期待の行は「空欄なら○」判定
        if what.endswith("空欄のままか"):
            f = f'=IF(D{r}="","",IF(OR(D{r}=0,D{r}="空欄"),"○","×"))'
        elif "氏名" in what or "組が更新" in what:
            f = f'=IF(D{r}="","",IF(OR(D{r}="OK",D{r}="○"),"○","要確認"))'
        else:
            f = f'=IF(D{r}="","",IF(D{r}=G{r},"○","×"))'
        jc = ws.cell(row=r, column=5, value=f)
        jc.fill = FILL_OUT; jc.border = BORDER; jc.alignment = CENTER
        jc.font = Font(name="游ゴシック", size=12, bold=True)
    ws.cell(row=5, column=7, value="（期待値）").font = F_NOTE
    ws.column_dimensions["G"].width = 12
    ws["A19"] = "※「氏名が書き換わっていないか」「組が更新されたか」の行は、目で見て問題なければ D列に ○ と入力してください。"
    ws["A19"].font = F_NOTE
    ws["A20"] = "※空欄を確認した行は D列に 0 または「空欄」と入力してください。"
    ws["A20"].font = F_NOTE

    # ---------- 4. 不具合記録 ----------
    ws = wb.create_sheet("④ 不具合記録")
    ws["A1"] = "不具合記録 ― ×が出たとき、エラーが出たときに書く"
    ws["A1"].font = F_TITLE
    ws["A2"] = "この4項目が埋まっていれば、修正版を最短で用意できます。エラー画面はスマホで撮って保存してください。"
    ws["A2"].font = F_NOTE
    head_row(ws, 4,
             ["No", "どのボタンか", "エラーメッセージ（画面の文言）", "直前にした操作", "使っていたデータ", "画面写真"],
             [5, 20, 40, 32, 18, 12])
    for i in range(10):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1).border = BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        for col in range(2, 7):
            c = ws.cell(row=r, column=col)
            c.fill = FILL_IN; c.border = BORDER; c.alignment = WRAP
        ws.row_dimensions[r].height = 34
    ws["A16"] = "【エラーが出たときの操作】エラー画面に「デバッグ」ボタンがあれば押して、黄色くなった行を撮影してください（原因が一発で特定できます）。"
    ws["A16"].font = F_NOTE

    # ファイルを開いた瞬間に数式を再計算させる（判定が空欄のままにならないように）
    wb.calculation.fullCalcOnLoad = True

    wb.save(path)
    print(f"  + {os.path.relpath(path, OUT)}")


# ==================================================================
# 同梱テキスト
# ==================================================================
README_ROOT = """＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝
　実機テストキット（自分用・2台のパソコンで検証する）
＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝

目的:
　学校へ持って行く前に、自分の手で「USBで渡したデータが
　マスターに正しく反映されるか」を確かめる。
　これが通れば、テスト導入の準備は完了です。

所要時間: 約2時間（初回のVBA組み込み30分＋テスト90分）

【当日の流れ】
　1.「03_記録用紙」の 2台PC実機テスト手順書.pdf を印刷する
　2. 手順書のとおり PC①（記録側）にVBAを組み込む ★初回だけ
　3. PC②（送信側）にデータを置き、USBでPC①へ運ぶ
　4. テストA〜Eを実行し、テスト結果記録ブック.xlsx に数値を記入
　5. 全部○になれば合格。×があれば「④不具合記録」に書いて連絡

【フォルダの中身】
　01_PC1_記録側/   … PC①に入れるもの（アシスタント・VBA・練習/検証データ）
　02_PC2_送信側/   … PC②に置くもの（学校に届くデータの想定）
　03_記録用紙/     … 手順書PDFと、結果を書き込むExcelブック

【大事なこと】
　・本物の生徒データは一切使いません（すべて架空データ）
　・失敗しても、このキットからコピーし直せば何度でもやり直せます
　・「壊れないか」を試すテストも含まれています（テストE）
"""

README_PC1 = """PC①（記録側）に入れるもの

このフォルダを丸ごと、PC①のデスクトップにコピーしてください。

1_アシスタント/     … 積立金入力アシスタント.xlsx と VBAモジュール12本
　　　　　　　　　　　★初回だけ、VBAを組み込んで .xlsm にする作業が必要
　　　　　　　　　　　（手順書の「準備2」を参照）
2_練習用データ/     … 架空80名。テストA〜Cで使う
3_検証用データ/     … 架空320名（実物と同じ規模）。テストDで使う

※このPCが「保管用パソコン」の役です。
"""

README_PC2 = """PC②（送信側）に置くもの

このフォルダを丸ごと、PC②のデスクトップにコピーしてください。

1_届いたデータ/     … 銀行から届いた振替結果・学校から届いた名簿の想定
2_PC1へ運ぶ箱/      … ここに入れたものをUSBでPC①へ運ぶ

※このPCが「データを受け取るパソコン」の役です。
※PC②にExcelが入っていなくても構いません（ファイルを運ぶだけ）。
"""


def main():
    if os.path.exists(KIT):
        shutil.rmtree(KIT)
    os.makedirs(KIT)

    write_text(os.path.join(KIT, "00_テストのはじめかた.txt"), README_ROOT)

    # ---- 01 PC1 記録側 ----
    d1 = os.path.join(KIT, "01_PC1_記録側")
    os.makedirs(d1)
    write_text(os.path.join(d1, "0_このフォルダについて.txt"), README_PC1)

    a = os.path.join(d1, "1_アシスタント")
    os.makedirs(a)
    copy("assistant/output/積立金入力アシスタント.xlsx", a)
    vba_dst = os.path.join(a, "VBAモジュール")
    os.makedirs(vba_dst)
    vba_src = os.path.join(REPO, "assistant", "vba")
    for name in sorted(os.listdir(vba_src)):
        if name.endswith(".bas"):
            shutil.copy2(os.path.join(vba_src, name), vba_dst)
    print("  + 01_PC1_記録側/1_アシスタント/VBAモジュール/(12本)")
    write_text(os.path.join(a, "★最初にやること.txt"),
               "【初回だけの作業・約30分】\r\n"
               "\r\n"
               "1. 積立金入力アシスタント.xlsx を開く\r\n"
               "2. Alt+F11 → ファイル → ファイルのインポート\r\n"
               "   → VBAモジュール フォルダの A00〜A11（12本）をすべて取り込む\r\n"
               "3. Alt+F11 で戻る → Alt+F8 →「初期設定」を実行\r\n"
               "   → メニューにボタン①〜⑮が並べば成功\r\n"
               "4. 名前を付けて保存 → ファイルの種類「Excelマクロ有効ブック(.xlsm)」\r\n"
               "\r\n"
               "この .xlsm が今後ずっと使う本体になります。\r\n"
               "（学校に渡すときも、この .xlsm をコピーするだけです）\r\n")

    p = os.path.join(d1, "2_練習用データ")
    os.makedirs(p)
    for f in ["練習用_令和X年度生積立金.xlsx", "練習用_口座マスター.xlsx",
              "練習用_振替結果.xlsx", "練習用_掲示用名簿.xlsx", "練習用_空のマスター.xlsx"]:
        copy(f"assistant/output/{f}", p)
    write_text(os.path.join(p, "このデータの期待値.txt"),
               "架空の80名（4クラス×20名）\r\n"
               "\r\n"
               "・①名簿解析 → 4クラス80名・全員一致\r\n"
               "・④支出（3500円・全員） → 対象80名／合計280,000円\r\n"
               "・⑪振替照合 → 読取80／振替済78／未納2（精算番号7・44）／不明0\r\n"
               "・⑤収入（76000円） → 入金78名／未納2名\r\n")

    v = os.path.join(d1, "3_検証用データ")
    os.makedirs(v)
    for f in ["検証用_令和X年度生積立金.xlsx", "検証用_口座マスター.xlsx",
              "検証用_振替結果.xlsx", "検証用_掲示用名簿.xlsx", "検証用_新年度名簿.xlsx"]:
        copy(f"assistant/output/{f}", v)
    write_text(os.path.join(v, "このデータの期待値.txt"),
               "架空の320名（8クラス×40名）― 実物と同じ規模\r\n"
               "\r\n"
               "・①名簿解析（検証用_新年度名簿） → 8クラス320名・全員一致\r\n"
               "・⑪振替照合（検証用_振替結果 321行）\r\n"
               "　 → 読取321／振替済315／未納5（精算番号7・44・159・241・312）／不明1\r\n"
               "　 ※不明1件は「見つからない口座があっても正しく検出できるか」を試すための\r\n"
               "　　 わざと仕込んだデータです。異常ではありません。\r\n")

    # ---- 02 PC2 送信側 ----
    d2 = os.path.join(KIT, "02_PC2_送信側")
    os.makedirs(d2)
    write_text(os.path.join(d2, "0_このフォルダについて.txt"), README_PC2)
    r1 = os.path.join(d2, "1_届いたデータ")
    os.makedirs(r1)
    copy("assistant/output/練習用_振替結果.xlsx", r1)
    copy("assistant/output/練習用_掲示用名簿.xlsx", r1)
    copy("assistant/output/検証用_振替結果.xlsx", r1)
    copy("assistant/output/検証用_新年度名簿.xlsx", r1)
    r2 = os.path.join(d2, "2_PC1へ運ぶ箱")
    os.makedirs(r2)
    write_text(os.path.join(r2, "使い方.txt"),
               "「1_届いたデータ」から、PC①へ渡すファイルをここにコピーし、\r\n"
               "USBメモリでPC①へ運びます。\r\n"
               "（学校では、この受け渡しを毎月行うことになります）\r\n")

    # ---- 03 記録用紙 ----
    d3 = os.path.join(KIT, "03_記録用紙")
    os.makedirs(d3)
    build_record_book(os.path.join(d3, "テスト結果記録ブック.xlsx"))
    copy("docs/05_開発・検証記録/2台PC実機テスト手順書.pdf", d3, required=False)
    copy("docs/01_学校向けマニュアル/動作確認チェックシート.pdf", d3, required=False)

    # ---- ZIP ----
    zip_path = os.path.join(OUT, "実機テストキット.zip")
    if os.path.exists(zip_path):
        os.remove(zip_path)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(KIT):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.relpath(fp, OUT))
    print(f"\n完成: {KIT}")
    print(f"完成: {zip_path} ({os.path.getsize(zip_path)/1024/1024:.1f} MB)")


if __name__ == "__main__":
    main()
