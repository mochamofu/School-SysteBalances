# -*- coding: utf-8 -*-
"""練習用のダミーファイルを生成するスクリプト

実物の「令和○年度生積立金」マスターと「掲示用名簿」に構造をそろえた、
架空の氏名だけで作った練習用ファイルを2つ生成する。
本番ファイルを触る前の練習・動作確認に使う。

生成物（assistant/output/）:
  練習用_令和X年度生積立金.xlsx … データ/精算書シート入りのマスター(縮小版だが行列配置は実物と同一)
  練習用_掲示用名簿.xlsx        … 4クラス×20名のクラス発表名簿(実物と同じブロック配置)
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(54)

# 架空の氏名パーツ（実在の個人とは無関係）
SEI = ["あおき", "いしだ", "うえの", "えんどう", "おかべ", "かねこ", "きたむら", "くどう",
       "こばやし", "さいとう", "しみず", "すずき", "せきぐち", "たなか", "つちや", "とみた",
       "なかがわ", "にしむら", "ぬまた", "のぐち", "はやし", "ひらの", "ふくだ", "ほんだ",
       "まつい", "みずの", "むらた", "もりた", "やまね", "よしおか"]
MEI = ["あおい", "いつき", "うた", "えま", "おうすけ", "かえで", "きよら", "くるみ",
       "けんと", "こはる", "さくと", "しおり", "すばる", "せな", "そうた", "たくみ",
       "ちひろ", "つむぎ", "ときお", "なぎさ", "にこ", "ねね", "のあ", "はると",
       "ひなた", "ふうか", "ほまれ", "まひろ", "みなと", "ゆいと"]

N_CLASS = 4
N_PER_CLASS = 20


def fake_names(n):
    names = set()
    while len(names) < n:
        names.add(f"{random.choice(SEI)}　{random.choice(MEI)}")
    return sorted(names)  # 五十音順 ＝ 名簿らしい並び


def build_roster(names, out):
    """実物の掲示用名簿と同じ「クラスごとのブロック配置」を再現する"""
    wb = Workbook()
    ws = wb.active
    ws.title = "掲示用"
    per = len(names) // N_CLASS
    for k in range(N_CLASS):
        base = 2 + k * 10          # ブロックの左端列（実物は10列ごと）
        hdr = ws.cell(row=2, column=base + 2, value=f"1年{k + 1}組")
        hdr.font = Font(bold=True, size=14)
        ws.cell(row=3, column=base + 1, value="担任")
        for i in range(per):
            ws.cell(row=4 + i, column=base + 1, value=i + 1)              # 番号
            ws.cell(row=4 + i, column=base + 4, value=names[k * per + i]) # 氏名
    wb.save(out)
    print("saved", out)


def build_master(names, out, filled=False):
    """実物マスターと同じ行列配置の「データ」「精算書」シートを持つ練習用ブック。
    filled=True なら氏名・組・番号と収入・支出のサンプル金額も入れる。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "データ"

    # --- 見出し（実物と同じ位置に同じ意味のものを置く）---
    ws["B3"] = "令和　年度入学生積立金会計　個人別管理票（練習用）"
    ws["B3"].font = Font(bold=True, size=14)
    ws["A5"] = "精算\n番号"; ws["B5"] = "生徒\n番号"; ws["C5"] = "入学年度"
    ws["D5"] = "学年"; ws["E5"] = "組"; ws["F5"] = "番号"; ws["G5"] = "氏名"
    ws["H5"] = "未納管理"; ws["J5"] = "収入の部"
    ws["BA6"] = "還付金\n合計"; ws["BB6"] = "計画徴収\n合計"; ws["BC6"] = "収入\n合計"
    ws["BE5"] = "支出の部"
    ws.cell(row=6, column=159, value="支出合計")
    ws.cell(row=5, column=160, value="残金")
    for hcell in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "BC6", "BA6", "BB6"]:
        ws[hcell].font = Font(bold=True)
        ws[hcell].alignment = Alignment(wrap_text=True, horizontal="center")
    # 支出No行（8行目）… No.1〜No.100
    for n in range(1, 101):
        ws.cell(row=8, column=56 + n, value=f"No.{n}")
    ws["BD6"] = 1

    # --- 生徒行（9〜329行）: 精算番号と集計数式（実物と同じ式）---
    for r in range(9, 330):
        ws.cell(row=r, column=1, value=r - 8)  # 精算番号
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF((BC{r}-BA{r})<$BC$4,"未納",""))')
        ws.cell(row=r, column=9, value=f'=IF(H{r}="","",BC{r}-BA{r}-$BC$4)')
        ws.cell(row=r, column=53, value=f"=SUM(AZ{r})")
        ws.cell(row=r, column=54, value=f"=SUM(M{r}:AZ{r})")
        ws.cell(row=r, column=55, value=f"=SUM(J{r}:AZ{r})+BA{r}")
        ws.cell(row=r, column=159, value=f"=SUM(BE{r}:FB{r})")
        ws.cell(row=r, column=160, value=f"=BC{r}-FC{r}")
    # 端数行(330)・合計行(331)
    ws.cell(row=330, column=1, value=322); ws.cell(row=330, column=2, value="端数")
    ws.cell(row=330, column=55, value="=SUM(J330:BA330)")
    ws.cell(row=330, column=159, value="=SUM(BE330:FB330)")
    ws.cell(row=330, column=160, value="=BC330-FC330")
    ws.cell(row=331, column=1, value=323); ws.cell(row=331, column=2, value="合計")
    ws.cell(row=331, column=55, value="=SUM(BC9:BC330)")
    ws.cell(row=331, column=159, value="=SUM(FC9:FC330)")
    ws.cell(row=331, column=160, value="=SUM(FD9:FD330)")

    if filled:
        # 生徒を登録済みの状態にする（クラス替え練習用）
        for i, name in enumerate(names):
            r = 9 + i
            ws.cell(row=r, column=3, value=7)               # 入学年度
            ws.cell(row=r, column=4, value=1)               # 学年
            ws.cell(row=r, column=5, value=i // N_PER_CLASS + 1)  # 組
            ws.cell(row=r, column=6, value=i % N_PER_CLASS + 1)   # 番号
            ws.cell(row=r, column=7, value=name)            # 氏名
        # 収入枠1（J列）: 積立金 76,000円 全員入金済みのサンプル
        ws.cell(row=6, column=10, value="学年積立金（練習サンプル）")
        for i in range(len(names)):
            ws.cell(row=9 + i, column=10, value=76000)

    # --- 精算書シート（実物の仕組みを簡略再現: J3に精算番号）---
    ps = wb.create_sheet("精算書")
    ps["H3"] = "精算番号"; ps["J3"] = 1
    ps["H7"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,4)&" 年"'
    ps["I7"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,5)&" 組"'
    ps["J7"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,6)&" 番"'
    ps["K7"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,7)&"　さん　保 護 者　殿"'
    ps["H10"] = "積　立　金　精　算　書（練習用）"
    ps["H21"] = "① 収入合計"; ps["N21"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,55)'
    ps["H25"] = "② 支出合計"; ps["N25"] = '=VLOOKUP($J$3,データ!$A$6:$FD$331,159)'
    ps["H27"] = "③ 差引返還額"; ps["N27"] = "=N21-N25"

    wb.save(out)
    print("saved", out)


def main():
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    names = fake_names(N_CLASS * N_PER_CLASS)
    build_roster(names, os.path.join(out_dir, "練習用_掲示用名簿.xlsx"))
    # 登録済みマスター（クラス替え・支出入力の練習用）
    build_master(names, os.path.join(out_dir, "練習用_令和X年度生積立金.xlsx"), filled=True)
    # 空のマスター（新入生登録の練習用）
    build_master(names, os.path.join(out_dir, "練習用_空のマスター.xlsx"), filled=False)


if __name__ == "__main__":
    main()
