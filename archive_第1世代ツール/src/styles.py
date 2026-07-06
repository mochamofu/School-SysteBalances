# -*- coding: utf-8 -*-
"""共通スタイルユーティリティ"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

COLORS = {
    "input":        "ADD8E6",
    "header":       "4472C4",
    "header_light": "BDD7EE",
    "header_mid":   "2E75B6",
    "green":        "70AD47",
    "green_light":  "E2EFDA",
    "orange":       "ED7D31",
    "orange_light": "FCE4D6",
    "red":          "FF0000",
    "red_light":    "FFE0E0",
    "purple":       "7030A0",
    "navy":         "203864",
    "yellow":       "FFFF00",
    "yellow_light": "FFF2CC",
    "gray":         "D9D9D9",
    "white":        "FFFFFF",
    "black":        "000000",
    "unmatched":    "FF6B6B",
    "matched":      "90EE90",
}


def fill(color_key):
    color = COLORS.get(color_key, color_key)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def font(bold=False, size=10, color="black", name="游ゴシック"):
    return Font(bold=bold, size=size, color=COLORS.get(color, color), name=name)


def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header(cell, text, color="header", font_color="white",
                 bold=True, size=10, wrap=True):
    cell.value = text
    cell.fill = fill(color)
    cell.font = font(bold=bold, size=size, color=font_color)
    cell.alignment = align("center", "center", wrap=wrap)
    cell.border = border()


def style_input(cell, color="input"):
    cell.fill = fill(color)
    cell.border = border()
    cell.alignment = align("left", "center")


def style_readonly(cell, value=None):
    if value is not None:
        cell.value = value
    cell.fill = fill("gray")
    cell.border = border()
    cell.alignment = align("center", "center")


def style_formula(cell, formula=None):
    if formula is not None:
        cell.value = formula
    cell.fill = fill("gray")
    cell.border = border()
    cell.alignment = align("right", "center")


def style_title(ws, text, row=1, col=1, merge_cols=8, size=14):
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = font(bold=True, size=size, color="header")
    cell.alignment = align("center", "center")
    if merge_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + merge_cols - 1)
    ws.row_dimensions[row].height = 24


def style_guide_row(ws, row, label, content, label_col=1, content_col=2):
    lc = ws.cell(row=row, column=label_col)
    lc.value = label
    lc.fill = fill("header_light")
    lc.font = font(bold=True, size=10)
    lc.border = border()
    lc.alignment = align("right", "center")

    cc = ws.cell(row=row, column=content_col)
    cc.value = content
    cc.font = font(size=10)
    cc.border = border()
    cc.alignment = align("left", "center", wrap=True)


def setup_sheet(ws, freeze_row=2, freeze_col=1, zoom=90, grid=True):
    ws.freeze_panes = ws.cell(row=freeze_row, column=freeze_col)
    ws.sheet_view.showGridLines = grid
    ws.sheet_view.zoomScale = zoom


def set_col_width(ws, col_widths):
    for col, width in col_widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = width


def set_row_height(ws, row_heights):
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height
