# -*- coding: utf-8 -*-
"""ツール② CoPilotプロンプトガイド生成"""

from openpyxl import Workbook
import sys, os

sys.path.insert(0, os.path.dirname(__file__))
from styles import (fill, font, border, align, style_header, style_title,
                    style_guide_row, setup_sheet, set_col_width, set_row_height, COLORS)

PROMPTS = [
    ("🌅 毎朝の確認",
     "昨日届いたメールで、本庁からの通知や期限付きの依頼があれば教えてください。",
     "毎朝、パソコンを開いたら最初に",
     "1日の始まりに習慣化すると通知見落としがゼロに近づく"),
    ("🔍 授業料の通知を探す",
     "授業料に関する本庁からの通知を、直近1ヶ月分、日付順にまとめてください。",
     "授業料関連の通知を探したいとき",
     "「授業料」の部分を好きなキーワードに変えてOK"),
    ("🔍 就学支援金の申請期限",
     "就学支援金の申請期限に関する通知を探して、対応手順を教えてください。",
     "就学支援金の申請前",
     "申請期限の通知を一発で特定できる"),
    ("🔍 給付型奨学金",
     "給付型奨学金に関する本庁からの通知を直近3ヶ月分、日付順にまとめてください。",
     "奨学金関連の手続きの前",
     "給付型奨学金の事務掲示板が別にある場合は追記で指示"),
    ("📬 室長メールを探す",
     "室長から転送されたメールで、入学選抜に関するものを一覧にしてください。",
     "室長からの重要メールを探したいとき",
     "室長の名前を入れるとより正確になる"),
    ("⏰ 今月の締め切り確認",
     "メールとSharePointの中から、今月末が期限の提出物や対応事項を一覧にしてください。",
     "月の前半（締切見落とし防止に）",
     "月末にこれを聞くと翌月の準備になる"),
    ("⏰ 今日のタスク確認",
     "今日が期限のタスクと、昨日届いた本庁通知を教えてください。",
     "毎朝の確認（詳細版）",
     "To Doリストと合わせて使うと最強"),
    ("📖 マニュアルを読む",
     "給付型奨学金の支出命令書を作成する手順を、SharePoint上のマニュアルを参照して教えてください。",
     "支出命令書を作るとき",
     "PDFを探す・開く・目次を探す手間がなくなる"),
    ("📊 週のまとめ",
     "今週のOutlookメール、Teams、SharePointの更新から、学次会計に関係するものを要約してください。",
     "金曜の午後（翌週の準備に）",
     "複数プラットフォームを一括横断検索"),
    ("📊 月次レビュー",
     "先月の本庁通知フォルダの内容を、カテゴリ別に整理してください。",
     "月初めの業務整理",
     "月ごとの対応記録になる"),
    ("📋 To Do 登録",
     "このメールから期限付きのタスクを抽出して、To Doリストの「学次業務」に期限日付きで登録してください。",
     "期限付きメールを受け取ったとき",
     "CoPilot Layer 3（To Do 自動生成）の活用"),
    ("⚠ うまく探せないとき",
     "メールだけでなくSharePointとTeamsも含めて、授業料に関する通知を探してください。",
     "「見つかりません」と言われたとき",
     "検索範囲を明示的に広げると精度UP"),
    ("🔑 キーワードを増やす",
     "授業料 就学支援金 申請期限 に関する本庁からの通知を直近1ヶ月でまとめてください。",
     "関係ない結果が出るとき",
     "キーワードを増やすと欲しい情報だけが出る"),
    ("📝 年度末の引き継ぎ準備",
     "今年度の学次会計に関するOutlookメール・SharePoint更新を要約して、引き継ぎ書のドラフトを作ってください。",
     "年度末3月",
     "前任者の記憶に頼らない引き継ぎが実現"),
]

OUTLOOK_RULES = [
    ("本庁通知",    "差出人が本庁ドメイン（例: @metro.ed.jp）のメール",
     "「ルールと通知」→「新しいルール」→「差出人がアドレスを含む場合」→フォルダ「本庁通知」に移動"),
    ("期限付き",    '件名に「○月○日まで」「期限」「締切」「提出」を含むメール',
     "「ルールと通知」→「新しいルール」→「件名に特定の語句が含まれる」→複数キーワードを「または」で設定"),
    ("室長転送",    "室長（学次担当室長）からのメール",
     "「送信者が特定の人物の場合」→室長の名前またはメアドを入力→フォルダ「室長転送」に移動"),
    ("授業料事務",  '件名に「授業料」「就学支援金」「奨学金」「徴収」を含むメール',
     "複数キーワードを「または」で設定→フォルダ「授業料事務」に移動"),
]


def build_prompt_sheet(ws):
    ws.tab_color = COLORS["header"]
    style_title(ws, "CoPilot プロンプトガイド ─ 学次会計担当者用（デスクに置いて使う）",
                row=1, merge_cols=4, size=13)

    note = "CoPilotは普通の日本語でOKです。黄色のプロンプト文をそのままコピペしてCoPilotに貼り付けてください。"
    ws.cell(row=2, column=1).value = note
    ws.cell(row=2, column=1).font = font(size=9)
    ws.cell(row=2, column=1).alignment = align("left","center",wrap=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 28

    for col, h in enumerate(["シーン","CoPilotへの質問文（そのままコピペ可）","使うタイミング","ポイント"], start=1):
        style_header(ws.cell(row=3, column=col), h, color="header", size=10)

    for i, (scene, prompt, timing, tip) in enumerate(PROMPTS, start=4):
        ws.cell(row=i, column=1).value = scene
        ws.cell(row=i, column=1).font = font(bold=True, size=9)
        ws.cell(row=i, column=1).fill = fill("header_light")
        ws.cell(row=i, column=1).border = border()
        ws.cell(row=i, column=1).alignment = align("center","center",wrap=True)

        ws.cell(row=i, column=2).value = prompt
        ws.cell(row=i, column=2).font = font(size=10, name="Consolas")
        ws.cell(row=i, column=2).fill = fill("yellow_light")
        ws.cell(row=i, column=2).border = border()
        ws.cell(row=i, column=2).alignment = align("left","center",wrap=True)

        for col, val in [(3, timing), (4, tip)]:
            ws.cell(row=i, column=col).value = val
            ws.cell(row=i, column=col).font = font(size=9)
            ws.cell(row=i, column=col).border = border()
            ws.cell(row=i, column=col).alignment = align("left","center",wrap=True)

        ws.row_dimensions[i].height = 32

    set_col_width(ws, {1:22, 2:58, 3:26, 4:36})
    set_row_height(ws, {1:22, 2:28, 3:20})
    setup_sheet(ws, freeze_row=4, freeze_col=1)


def build_outlook_rules_sheet(ws):
    ws.tab_color = COLORS["green"]
    style_title(ws, "Outlook 自動分類ルール設定手順 (Layer 2)", row=1, merge_cols=3, size=13)

    note = "このルールを設定するとCoPilotの検索精度が大幅に向上します。Outlookで一度設定すれば以後は自動で動きます。"
    ws.cell(row=2, column=1).value = note
    ws.cell(row=2, column=1).font = font(size=9)
    ws.cell(row=2, column=1).alignment = align("left","center",wrap=True)
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 28

    for col, h in enumerate(["フォルダ名","ルール条件","設定手順"], start=1):
        style_header(ws.cell(row=3, column=col), h, color="green", font_color="white")

    for i, (folder, condition, steps) in enumerate(OUTLOOK_RULES, start=4):
        ws.cell(row=i, column=1).value = folder
        ws.cell(row=i, column=1).font = font(bold=True, size=10)
        ws.cell(row=i, column=1).fill = fill("header_light")
        ws.cell(row=i, column=1).border = border()
        ws.cell(row=i, column=1).alignment = align("center","center")

        for col, val in [(2, condition), (3, steps)]:
            ws.cell(row=i, column=col).value = val
            ws.cell(row=i, column=col).font = font(size=9)
            ws.cell(row=i, column=col).border = border()
            ws.cell(row=i, column=col).alignment = align("left","center",wrap=True)

        ws.row_dimensions[i].height = 44

    # 設定後の活用例
    ws.cell(row=9, column=1).value = "設定後のCoPilot活用例"
    ws.cell(row=9, column=1).font = font(bold=True, size=10, color="header")
    ws.merge_cells("A9:C9")
    for j, text in enumerate([
        "「本庁通知フォルダの中から就学支援金に関するメールを探してください」",
        "「期限付きフォルダの未読メールを期限日順にまとめてください」",
        "「室長転送フォルダの今週のメールを要約してください」",
    ], start=10):
        c = ws.cell(row=j, column=1)
        c.value = text
        c.font = font(size=10, name="Consolas")
        c.fill = fill("yellow_light")
        c.border = border()
        c.alignment = align("left","center")
        ws.merge_cells(f"A{j}:C{j}")
        ws.row_dimensions[j].height = 20

    set_col_width(ws, {1:16, 2:40, 3:55})
    set_row_height(ws, {1:22, 2:28, 3:20})
    setup_sheet(ws, freeze_row=4)


def build_calendar_sheet(ws):
    ws.tab_color = COLORS["purple"]
    style_title(ws, "CoPilot 活用 ─ 推奨運用サイクル", row=1, merge_cols=4, size=13)

    for col, h in enumerate(["頻度","アクション","CoPilotプロンプト","効果"], start=1):
        style_header(ws.cell(row=2, column=col), h, color="header")

    calendar = [
        ("毎朝","今日やるべきことの確認",
         "今日が期限のタスクと、昨日届いた本庁通知を教えてください",
         "通知見落とし防止、1日の優先順位が明確になる"),
        ("毎週金曜","週次の業務棚卸し",
         "今週のOutlookとSharePointの更新で、学次会計に関係するものを要約してください",
         "翌週への申し送り作成"),
        ("月初め","月次レビュー",
         "先月の本庁通知フォルダの内容を、カテゴリ別に整理してください",
         "月次の対応記録、引き継ぎ資料になる"),
        ("月の前半","今月の締切確認",
         "メールとSharePointの中から、今月末が期限の提出物や対応事項を一覧にしてください",
         "締切見落とし防止"),
        ("年度末3月","年度の引き継ぎ準備",
         "今年度の学次会計に関するOutlookメール・SharePoint更新を要約して、引き継ぎ書のドラフトを作ってください",
         "引き継ぎ書作成の負担が大幅減"),
        ("随時","特定通知の検索",
         "就学支援金の申請期限に関する通知を探して、対応手順を教えてください",
         "必要なときにいつでも即座に情報取得"),
    ]
    for i, (freq, action, prompt, effect) in enumerate(calendar, start=3):
        ws.cell(row=i, column=1).value = freq
        ws.cell(row=i, column=1).fill = fill("header_light")
        ws.cell(row=i, column=1).font = font(bold=True, size=10)
        ws.cell(row=i, column=1).border = border()
        ws.cell(row=i, column=1).alignment = align("center","center")

        ws.cell(row=i, column=2).value = action
        ws.cell(row=i, column=2).font = font(size=10)
        ws.cell(row=i, column=2).border = border()

        ws.cell(row=i, column=3).value = prompt
        ws.cell(row=i, column=3).font = font(size=10, name="Consolas")
        ws.cell(row=i, column=3).fill = fill("yellow_light")
        ws.cell(row=i, column=3).border = border()
        ws.cell(row=i, column=3).alignment = align("left","center",wrap=True)

        ws.cell(row=i, column=4).value = effect
        ws.cell(row=i, column=4).font = font(size=9)
        ws.cell(row=i, column=4).border = border()
        ws.cell(row=i, column=4).alignment = align("left","center",wrap=True)

        ws.row_dimensions[i].height = 28

    set_col_width(ws, {1:12, 2:22, 3:55, 4:35})
    set_row_height(ws, {1:22, 2:20})
    setup_sheet(ws, freeze_row=3)


def create_workbook(output_path):
    wb = Workbook()
    ws_prompt = wb.active
    ws_prompt.title = "プロンプトテンプレート"
    build_prompt_sheet(ws_prompt)
    build_outlook_rules_sheet(wb.create_sheet("Outlookルール設定"))
    build_calendar_sheet(wb.create_sheet("運用カレンダー"))
    wb.save(output_path)
    print(f"  ✓ CoPilotプロンプトガイド_学次会計用.xlsx → {output_path}")


if __name__ == "__main__":
    os.makedirs("../output", exist_ok=True)
    create_workbook("../output/CoPilotプロンプトガイド_学次会計用.xlsx")
